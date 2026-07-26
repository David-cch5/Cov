#!/usr/bin/env python3
"""
Paragraph-level breakdown of covenant versions.

Reads Covenant_Matrix/covenant_matrix.json (from cov_version_matrix.py) and:
 1. aligns paragraphs across versions into SLOTS (same paragraph, possibly reworded)
 2. within a slot, groups texts into VARIANTS (v1, v2, ... - different language;
    OCR noise does NOT create a new variant)
 3. rewrites covenant_matrix.xlsx with 5 sheets:
      Versions            one row per version (summary + full text chunks)
      CovID_Map           one row per document
      VersionParagraph_Map  LONG: VersionID, Position, SlotID, VariantID  <- FileMaker find table
      Paragraph_Variants  one row per variant: SlotID, VariantID, UsedByCount, Text
      Versions_Paragraphs WIDE: VersionID + one column per shared slot, cell = variant id
 4. rewrites covenant_matrix.html with paragraph variant badges + side-by-side
    version compare with word-level diff
 5. writes paragraphs.json

Usage: python3 cov_paragraph_matrix.py
Thresholds: SLOT_SIM couples reworded paragraphs into one slot;
            VARIANT_SIM separates real language changes from OCR noise.
"""
import os, re, json, time, datetime
import numpy as np
from collections import Counter, defaultdict
from difflib import SequenceMatcher

SRC = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SRC, "Covenant_Matrix")
SLOT_SIM = 0.60
VARIANT_SIM = 0.78   # difflib ratio; OCR noise scores ~0.81+, real rewording ~0.71
CHUNK = 30000

t0 = time.time()
data = json.load(open(os.path.join(OUT, "covenant_matrix.json"), encoding="utf-8"))
versions = data["versions"]
mapping = data["covid_map"]
stats = data["stats"]

# ---------------------------------------------------------------- collect paragraphs
paras, owner = [], []           # text, (version_index, position)
for vi, v in enumerate(versions):
    for pi, p in enumerate(v["paragraphs"]):
        paras.append(p)
        owner.append((vi, pi))
print(f"{len(paras)} paragraphs from {len(versions)} versions")

def norm(t):
    t = t.lower()
    t = re.sub(r"[0-9]+", "#", t)
    t = re.sub(r"[^a-z# ]+", " ", t)
    return re.sub(r"\s+", " ", t).strip()

norms = [norm(p) for p in paras]

# ---------------------------------------------------------------- slot clustering
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.neighbors import NearestNeighbors
from scipy.sparse import csr_matrix
from scipy.sparse.csgraph import connected_components

import pickle
CKPT = "/tmp/para_ckpt.pkl"
CKPT2 = "/tmp/para_ckpt2.pkl"
if os.path.exists(CKPT2):
    X, slot_lab = pickle.load(open(CKPT2, "rb"))
    REFINED = True
    print(f"refined checkpoint loaded {time.time()-t0:.0f}s")
elif os.path.exists(CKPT):
    X, slot_lab = pickle.load(open(CKPT, "rb"))
    REFINED = False
    print(f"checkpoint loaded {time.time()-t0:.0f}s")
else:
    REFINED = False
    vec = TfidfVectorizer(analyzer="char_wb", ngram_range=(4, 4), max_features=200000,
                          sublinear_tf=True)
    X = vec.fit_transform(norms)
    print(f"tfidf {time.time()-t0:.0f}s")
    nn = NearestNeighbors(radius=1 - SLOT_SIM, metric="cosine").fit(X)
    G = nn.radius_neighbors_graph(X, mode="connectivity")
    print(f"neighbors {time.time()-t0:.0f}s")
    ncomp, slot_lab = connected_components(G, directed=False)
    pickle.dump((X, slot_lab), open(CKPT, "wb"))
print(f"{len(set(slot_lab))} components  {time.time()-t0:.0f}s")

# refine: connected components chain unrelated paragraphs together, so re-cluster
# each component with average linkage (no chaining) at the same threshold
if not REFINED:
    from sklearn.preprocessing import normalize as _l2
    from scipy.cluster.hierarchy import linkage as _lk, fcluster as _fc
    from scipy.spatial.distance import squareform as _sq
    _Xn = _l2(X)
    comp_members = defaultdict(list)
    for i, l in enumerate(slot_lab):
        comp_members[l].append(i)
    next_lab = 0
    refined = np.empty(len(slot_lab), dtype=np.int64)
    for l, members in comp_members.items():
        if len(members) <= 2:
            for i in members:
                refined[i] = next_lab
            next_lab += 1
            continue
        Xs = _Xn[members]
        Dm = 1 - (Xs @ Xs.T).toarray()
        np.fill_diagonal(Dm, 0)
        Dm = np.clip(Dm, 0, None)
        sub = _fc(_lk(_sq(Dm, checks=False), method="average"),
                  1 - SLOT_SIM, criterion="distance")
        for i, s in zip(members, sub):
            refined[i] = next_lab + s
        next_lab += int(sub.max()) + 1
    slot_lab = refined
    pickle.dump((X, slot_lab), open(CKPT2, "wb"))
    print(f"{len(set(slot_lab))} slots after refinement  {time.time()-t0:.0f}s")

# order slots by median position of their paragraphs
slot_members = defaultdict(list)
for i, l in enumerate(slot_lab):
    slot_members[l].append(i)
slot_order = sorted(slot_members,
                    key=lambda l: (np.median([owner[i][1] for i in slot_members[l]]),
                                   -len(slot_members[l])))
slot_id = {l: f"P{r+1:03d}" for r, l in enumerate(slot_order)}

# ---------------------------------------------------------------- variants within slot
# greedy leader clustering on the same char-4gram tfidf vectors (fast, OCR-tolerant)
from sklearn.preprocessing import normalize as l2norm
import sys
Xn = l2norm(X)
VCKPT = "/tmp/para_var.pkl"
if os.path.exists(VCKPT):
    variant_of, variants, done_slots = pickle.load(open(VCKPT, "rb"))
    print(f"variant checkpoint: {len(done_slots)} slots done")
else:
    variant_of, variants, done_slots = {}, {}, set()
for l in sorted(slot_members):
    if l in done_slots:
        continue
    if time.time() - t0 > 36:
        pickle.dump((variant_of, variants, done_slots), open(VCKPT, "wb"))
        print(f"PARTIAL: {len(done_slots)}/{len(slot_members)} slots — rerun to continue")
        sys.exit(1)
    members = slot_members[l]
    sid = slot_id[l]
    members = sorted(members, key=lambda i: -len(norms[i]))
    if len(members) == 1:
        i = members[0]
        variants[(sid, 1)] = {"rep": i, "members": [i]}
        variant_of[i] = (sid, 1)
        done_slots.add(l)
        continue
    Xs = Xn[members]
    S = (Xs @ Xs.T).toarray()          # pairwise cosine within slot (prefilter)
    big = len(members) > 200           # huge slots (mostly OCR junk): cosine only
    lvo, lvar = {}, {}                 # local results; committed atomically per slot
    leaders = []                       # local indices
    for li, i in enumerate(members):
        vn = None
        if leaders:
            order = np.argsort(-S[li, leaders])[:15]
            for oi in order:
                lj = leaders[oi]
                c = S[li, lj]
                if c >= 0.95 or (big and c >= 0.80):
                    vn = oi + 1; break
                if c < 0.55 or big:
                    break
                a, b = norms[i][:1500], norms[members[lj]][:1500]
                sm = SequenceMatcher(None, a, b)
                if sm.quick_ratio() >= VARIANT_SIM and sm.ratio() >= VARIANT_SIM:
                    vn = oi + 1; break
        if vn is None:
            leaders.append(li)
            vn = len(leaders)
            lvar[(sid, vn)] = {"rep": i, "members": []}
        lvar[(sid, vn)]["members"].append(i)
        lvo[i] = (sid, vn)
    variants.update(lvar)
    variant_of.update(lvo)
    done_slots.add(l)
print(f"variants {time.time()-t0:.0f}s: {len(variants)} total")

# renumber variants by usage (v1 = most used, count = distinct versions)
vcount = {}
for (sid, vn), d in variants.items():
    vcount[(sid, vn)] = len({owner[i][0] for i in d["members"]})
new_no = {}
for sid in set(s for s, _ in variants):
    vs = sorted([k for k in variants if k[0] == sid],
                key=lambda k: (-vcount[k], k[1]))
    for newn, k in enumerate(vs, 1):
        new_no[k] = newn
variant_of = {i: (s, new_no[(s, v)]) for i, (s, v) in variant_of.items()}
variants = {(s, new_no[(s, v)]): d for (s, v), d in variants.items()}

# attach to versions
for vi, v in enumerate(versions):
    v["para_slots"] = [None] * len(v["paragraphs"])
for i, (vi, pi) in enumerate(owner):
    sid, vn = variant_of[i]
    versions[vi]["para_slots"][pi] = [sid, vn]

# usage maps
slot_versions = defaultdict(set)      # slotID -> set(version ids)
variant_versions = defaultdict(set)   # (sid,vn) -> set(version ids)
for i, (vi, pi) in enumerate(owner):
    sid, vn = variant_of[i]
    slot_versions[sid].add(versions[vi]["version_id"])
    variant_versions[(sid, vn)].add(versions[vi]["version_id"])

shared_slots = [s for s in sorted(slot_versions) if len(slot_versions[s]) >= 2]
print(f"slots used by >=2 versions: {len(shared_slots)}")

# ---------------------------------------------------------------- xlsx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
wb = Workbook()
bold = Font(name="Arial", bold=True, color="FFFFFF")
fill = PatternFill("solid", fgColor="1F4E79")
arial = Font(name="Arial")

def style(ws, widths=None):
    for r in ws.iter_rows():
        for c in r: c.font = arial
    for c in ws[1]: c.font = bold; c.fill = fill
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

def xclean(s):
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", s)

# Versions
ws = wb.active; ws.title = "Versions"
max_chunks = max((v["char_count"] // CHUNK) + 1 for v in versions)
ws.append(["VersionID", "DocCount", "PageCount", "ParagraphCount", "SampleCovID",
           "Cohesion", "OCR_Used"] + [f"VersionText_{i+1}" for i in range(max_chunks)])
for v in versions:
    text = xclean("\n\n".join(v["paragraphs"]))
    chunks = [text[i:i+CHUNK] for i in range(0, len(text), CHUNK)] or [""]
    ws.append([v["version_id"], v["doc_count"], v["page_count"], len(v["paragraphs"]),
               v["sample_covid"], v["cohesion"], "Yes" if v["ocr_used"] else "No"]
              + chunks + [""] * (max_chunks - len(chunks)))
style(ws, [10, 9, 9, 12, 11, 9, 9])

# CovID_Map
ws2 = wb.create_sheet("CovID_Map")
ws2.append(["CovID", "VersionID", "Filename", "RelativePath", "Pages", "OCR_Used",
            "SimilarityToVersion"])
for m in mapping:
    ws2.append([m["covid"], m["version_id"], m["filename"], m["relpath"],
                m["pages"], "Yes" if m["ocr"] else "No", m["sim_to_rep"]])
style(ws2, [10, 10, 30, 40, 8, 10, 18])

# VersionParagraph_Map (long — the FileMaker find table)
ws3 = wb.create_sheet("VersionParagraph_Map")
ws3.append(["VersionID", "Position", "SlotID", "VariantID"])
for v in versions:
    for pi, sv in enumerate(v["para_slots"], 1):
        sid, vn = sv
        ws3.append([v["version_id"], pi, sid, f"{sid}_v{vn}"])
style(ws3, [10, 9, 9, 12])

# Paragraph_Variants (SimilarityToV1 near 1.0 on a v2+ row = likely OCR artifact,
# not a real language change)
ws4 = wb.create_sheet("Paragraph_Variants")
ws4.append(["SlotID", "VariantID", "VariantNo", "UsedByVersionCount",
            "SimilarityToV1", "UsedByVersions", "Text"])
v1rep = {sid: d["rep"] for (sid, vn), d in variants.items() if vn == 1}
for (sid, vn), d in sorted(variants.items()):
    used = sorted(variant_versions[(sid, vn)])
    if vn == 1:
        s2v1 = 1.0
    else:
        s2v1 = round(SequenceMatcher(None, norms[d["rep"]][:1500],
                                     norms[v1rep[sid]][:1500]).ratio(), 3)
    ws4.append([sid, f"{sid}_v{vn}", vn, len(used), s2v1,
                ", ".join(used[:120]), xclean(paras[d["rep"]])[:32000]])
style(ws4, [9, 12, 10, 18, 14, 40, 80])

# Versions_Paragraphs (wide)
ws5 = wb.create_sheet("Versions_Paragraphs")
ws5.append(["VersionID", "DocCount"] + shared_slots + ["UniqueSlotCount"])
col_of = {s: i for i, s in enumerate(shared_slots)}
for v in versions:
    row = [""] * len(shared_slots)
    uniq = 0
    for sv in v["para_slots"]:
        sid, vn = sv
        if sid in col_of:
            cur = row[col_of[sid]]
            tag = f"v{vn}"
            row[col_of[sid]] = tag if not cur else (cur if tag in cur.split("+") else cur + "+" + tag)
        else:
            uniq += 1
    ws5.append([v["version_id"], v["doc_count"]] + row + [uniq])
style(ws5, [10, 9] + [7] * len(shared_slots) + [14])
wb.save(os.path.join(OUT, "covenant_matrix.xlsx"))
print(f"xlsx {time.time()-t0:.0f}s")

# ---------------------------------------------------------------- json
json.dump({"stats": stats,
           "versions": [{k: v[k] for k in ("version_id", "doc_count", "sample_covid",
                                            "cohesion", "covids", "paragraphs",
                                            "para_slots")} for v in versions],
           "covid_map": mapping,
           "paragraph_variants": [
               {"slot": sid, "variant": f"{sid}_v{vn}",
                "used_by": sorted(variant_versions[(sid, vn)]),
                "text": paras[d["rep"]]}
               for (sid, vn), d in sorted(variants.items())]},
          open(os.path.join(OUT, "paragraphs.json"), "w", encoding="utf-8"), indent=1)

# ---------------------------------------------------------------- html
hdata = {"stats": dict(stats, slots=len(slot_versions), sharedSlots=len(shared_slots),
                       variants=len(variants)),
         "versions": [{"id": v["version_id"], "docCount": v["doc_count"],
                       "pages": v["page_count"], "cohesion": v["cohesion"],
                       "sample": v["sample_covid"], "covids": v["covids"],
                       "paragraphs": v["paragraphs"],
                       "slots": [f"{s}_v{n}" for s, n in v["para_slots"]]}
                      for v in versions]}
payload = json.dumps(hdata).replace("</", "<\\/")
tpl = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Covenant Version Matrix</title>
<style>
 :root{--blue:#1F4E79;--lt:#eef3f8;--hl:#ffe066;--add:#d4f7d4;--del:#ffd9d9}
 *{box-sizing:border-box}
 body{font-family:-apple-system,Segoe UI,Arial,sans-serif;margin:0;background:#f5f6f8;color:#222}
 header{background:var(--blue);color:#fff;padding:14px 28px}
 header h1{margin:0 0 6px;font-size:20px}
 .stats{display:flex;gap:24px;font-size:12.5px;opacity:.92;flex-wrap:wrap}
 .stats b{font-size:16px;display:block}
 .bar{position:sticky;top:0;background:#fff;padding:10px 28px;box-shadow:0 1px 4px rgba(0,0,0,.12);z-index:5;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
 #q{flex:1;min-width:240px;max-width:520px;padding:8px 13px;font-size:14.5px;border:1.5px solid #bbb;border-radius:8px}
 #hits{font-size:13px;color:#555}
 label.f{font-size:13px;color:#444;display:flex;gap:5px;align-items:center}
 #cmpbtn{padding:7px 14px;border:1.5px solid var(--blue);background:#fff;color:var(--blue);border-radius:8px;cursor:pointer;font-size:13.5px}
 #cmpbtn:disabled{opacity:.4;cursor:default}
 main{padding:16px 28px;max-width:1250px;margin:0 auto}
 .ver{background:#fff;border:1px solid #dde3ea;border-radius:10px;margin-bottom:11px;overflow:hidden}
 .vhead{display:flex;align-items:center;gap:12px;padding:10px 15px;cursor:pointer;background:var(--lt)}
 .vhead:hover{background:#e2ebf5}
 .vid{font-weight:700;color:var(--blue);font-size:16px;min-width:52px}
 .meta{font-size:12.5px;color:#555}
 .cmp{margin-left:auto;font-size:12px;display:flex;gap:5px;align-items:center;color:#444}
 .chips{padding:7px 15px;border-top:1px solid #e8edf2;display:flex;flex-wrap:wrap;gap:4px;max-height:100px;overflow:auto}
 .chip{background:var(--lt);border:1px solid #c9d7e6;border-radius:12px;padding:1px 8px;font-size:11.5px;font-family:ui-monospace,monospace}
 .chip.hit{background:var(--hl);border-color:#d4b106}
 .body{display:none;border-top:1px solid #e8edf2;padding:11px 15px;max-height:520px;overflow:auto}
 .ver.open .body{display:block}
 .para{display:flex;gap:10px;padding:4px 0;font-size:13.5px;line-height:1.5}
 .pn{color:#fff;background:#8fa8c2;border-radius:4px;font-family:ui-monospace,monospace;font-size:10.5px;min-width:74px;text-align:center;flex-shrink:0;align-self:flex-start;padding:2px 4px;cursor:pointer}
 .pn.vhi{background:#c0392b}
 mark{background:var(--hl);padding:0 1px}
 .arrow{transition:.2s;color:#888}
 .ver.open .arrow{transform:rotate(90deg)}
 .hidden{display:none}
 #overlay{display:none;position:fixed;inset:0;background:rgba(20,30,45,.55);z-index:20}
 #panel{position:fixed;inset:3% 4%;background:#fff;border-radius:12px;z-index:21;display:none;flex-direction:column;overflow:hidden}
 #phead{background:var(--blue);color:#fff;padding:12px 20px;display:flex;gap:16px;align-items:center}
 #phead b{font-size:16px}
 #pclose{margin-left:auto;background:none;border:1px solid #fff;color:#fff;border-radius:6px;padding:4px 12px;cursor:pointer}
 #pbody{overflow:auto;padding:14px 20px;flex:1}
 .drow{display:grid;grid-template-columns:90px 1fr 1fr;gap:10px;border-bottom:1px solid #eef1f5;padding:7px 0;font-size:13px;line-height:1.5}
 .drow.same{color:#999;font-size:12px}
 .drow .slot{font-family:ui-monospace,monospace;font-size:11px;color:#555}
 .drow ins{background:var(--add);text-decoration:none}
 .drow del{background:var(--del);text-decoration:none}
 .dhead{position:sticky;top:0;background:#fff;font-weight:700;border-bottom:2px solid var(--blue)}
 label.g{font-size:12.5px;color:#eee;display:flex;gap:5px;align-items:center}
</style></head><body>
<header><h1>Covenant Version Matrix — paragraph level</h1>
<div class="stats">
 <span><b id=s1></b>documents</span><span><b id=s2></b>versions</span>
 <span><b id=s3></b>paragraph slots</span><span><b id=s4></b>shared slots</span>
 <span><b id=s5></b>language variants</span><span><b id=s6></b>generated</span>
</div></header>
<div class="bar"><input id="q" placeholder="Search text, CovID, or paragraph ID (e.g. 4417, transfer fee, P012_v2)…" autofocus>
<label class="f"><input type="checkbox" id="multi" checked> hide one-off versions</label>
<button id="cmpbtn" disabled>Compare (pick 2)</button>
<span id="hits"></span></div>
<main id="list"></main>
<div id="overlay"></div>
<div id="panel"><div id="phead"><b id=ptitle></b>
<label class="g"><input type="checkbox" id="hidesame" checked> hide identical paragraphs</label>
<button id="pclose">Close</button></div><div id="pbody"></div></div>
<script>
const DATA = __DATA__;
const esc=s=>s.replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));
s1.textContent=DATA.stats.docs;s2.textContent=DATA.stats.versions;
s3.textContent=DATA.stats.slots;s4.textContent=DATA.stats.sharedSlots;
s5.textContent=DATA.stats.variants;s6.textContent=DATA.stats.date;
const list=document.getElementById('list'),els={},sel=new Set();
DATA.versions.forEach(v=>{
 const d=document.createElement('div');d.className='ver';d.dataset.id=v.id;
 d.innerHTML=`<div class="vhead"><span class="vid">${v.id}</span>
  <span class="meta">${v.docCount} doc${v.docCount!==1?'s':''} · ${v.pages}p · ${v.paragraphs.length} ¶ · sample ${v.sample}</span>
  <label class="cmp" onclick="event.stopPropagation()"><input type="checkbox" data-v="${v.id}" class="selbox">compare</label>
  <span class="arrow">▶</span></div>
  <div class="chips">${v.covids.slice(0,400).map(c=>`<span class="chip" data-c="${c}">${c}</span>`).join('')}</div>
  <div class="body"></div>`;
 d.querySelector('.vhead').onclick=()=>{d.classList.toggle('open');if(d.classList.contains('open'))render(v,d,curQ());};
 els[v.id]=d;list.appendChild(d);
});
document.querySelectorAll('.selbox').forEach(b=>b.addEventListener('change',()=>{
 if(b.checked)sel.add(b.dataset.v);else sel.delete(b.dataset.v);
 if(sel.size>2){b.checked=false;sel.delete(b.dataset.v);return;}
 const btn=document.getElementById('cmpbtn');
 btn.disabled=sel.size!==2;
 btn.textContent=sel.size===2?`Compare ${[...sel].join(' vs ')}`:'Compare (pick 2)';
}));
function curQ(){return document.getElementById('q').value.trim().toLowerCase()}
function render(v,el,s){
 const rx=s?new RegExp('('+s.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+')','gi'):null;
 let out=[];
 v.paragraphs.forEach((p,i)=>{
  const tag=v.slots[i];
  const tagHit=s&&tag.toLowerCase().includes(s);
  if(s&&!p.toLowerCase().includes(s)&&!tagHit)return;
  let h=esc(p);if(rx&&!tagHit)h=h.replace(rx,'<mark>$1</mark>');
  out.push(`<div class="para"><span class="pn${tagHit?' vhi':''}" title="paragraph slot / language variant">${tag}</span><span>${h}</span></div>`);
 });
 if(s&&!out.length)out.push('<div class="para"><span></span><span style="color:#999">CovID match only</span></div>');
 el.querySelector('.body').innerHTML=out.join('');
}
const q=document.getElementById('q'),hits=document.getElementById('hits'),multi=document.getElementById('multi');
let t;q.addEventListener('input',()=>{clearTimeout(t);t=setTimeout(search,250)});
multi.addEventListener('change',search);
function search(){
 const s=curQ();let shown=0,pm=0;
 DATA.versions.forEach(v=>{
  const el=els[v.id];
  el.querySelectorAll('.chip.hit').forEach(c=>c.classList.remove('hit'));
  if(!s){const hide=multi.checked&&v.docCount===1;
   el.classList.toggle('hidden',hide);el.classList.remove('open');
   if(!hide)shown++;return;}
  const cov=v.covids.filter(c=>c.includes(s));
  let tp=0;
  v.paragraphs.forEach((p,i)=>{if(p.toLowerCase().includes(s)||v.slots[i].toLowerCase().includes(s)){tp++;pm++;}});
  const match=tp>0||cov.length>0;
  el.classList.toggle('hidden',!match);
  if(match){shown++;el.classList.add('open');render(v,el,s);
   cov.forEach(c=>el.querySelector(`.chip[data-c="${c}"]`)?.classList.add('hit'));}
 });
 hits.textContent=s?`${shown} versions, ${pm} matching paragraphs`:'';
}
search();
// ---------- compare ----------
function wdiff(a,b){ // word LCS diff -> [ [tag,text] ]
 const A=a.split(/\s+/),B=b.split(/\s+/),n=A.length,m=B.length;
 if(n*m>4000000)return [['rep',a,b]];
 const L=Array.from({length:n+1},()=>new Int32Array(m+1));
 for(let i=n-1;i>=0;i--)for(let j=m-1;j>=0;j--)
  L[i][j]=A[i]===B[j]?L[i+1][j+1]+1:Math.max(L[i+1][j],L[i][j+1]);
 const out=[];let i=0,j=0;
 while(i<n&&j<m){
  if(A[i]===B[j]){out.push(['eq',A[i]]);i++;j++;}
  else if(L[i+1][j]>=L[i][j+1]){out.push(['del',A[i]]);i++;}
  else{out.push(['ins',B[j]]);j++;}
 }
 while(i<n)out.push(['del',A[i++]]);
 while(j<m)out.push(['ins',B[j++]]);
 return out;
}
function renderDiff(a,b){
 const ops=wdiff(a,b);
 if(ops.length===1&&ops[0][0]==='rep')return [esc(a),esc(b)];
 let la=[],lb=[];
 ops.forEach(([t,w])=>{
  if(t==='eq'){la.push(esc(w));lb.push(esc(w));}
  else if(t==='del')la.push('<del>'+esc(w)+'</del>');
  else lb.push('<ins>'+esc(w)+'</ins>');
 });
 return [la.join(' '),lb.join(' ')];
}
document.getElementById('cmpbtn').onclick=()=>{
 const [ida,idb]=[...sel];
 const va=DATA.versions.find(v=>v.id===ida),vb=DATA.versions.find(v=>v.id===idb);
 const mapa={},mapb={};
 va.slots.forEach((s,i)=>{const k=s.split('_')[0];(mapa[k]=mapa[k]||[]).push(i);});
 vb.slots.forEach((s,i)=>{const k=s.split('_')[0];(mapb[k]=mapb[k]||[]).push(i);});
 const keys=[...new Set([...va.slots,...vb.slots].map(s=>s.split('_')[0]))];
 keys.sort((x,y)=>{
  const px=mapa[x]?mapa[x][0]:1e9, py=mapa[y]?mapa[y][0]:1e9;
  const qx=mapb[x]?mapb[x][0]:1e9, qy=mapb[y]?mapb[y][0]:1e9;
  return Math.min(px,qx)-Math.min(py,qy)||px-py;
 });
 let rows=[`<div class="drow dhead"><span class="slot">Slot</span><span>${va.id}</span><span>${vb.id}</span></div>`];
 let same=0,diff=0;
 keys.forEach(k=>{
  const ia=mapa[k]||[],ib=mapb[k]||[];
  const n=Math.max(ia.length,ib.length);
  for(let x=0;x<n;x++){
   const pa=ia[x]!==undefined?va.paragraphs[ia[x]]:null;
   const pb=ib[x]!==undefined?vb.paragraphs[ib[x]]:null;
   const ta=ia[x]!==undefined?va.slots[ia[x]]:'—';
   const tb=ib[x]!==undefined?vb.slots[ib[x]]:'—';
   if(pa!==null&&pb!==null&&ta===tb){same++;
    rows.push(`<div class="drow same hideable"><span class="slot">${ta}</span><span colspan=2>identical language (${ta})</span><span></span></div>`);
   }else{diff++;
    const [ha,hb]=pa!==null&&pb!==null?renderDiff(pa,pb):[pa?esc(pa):'<i style="color:#bbb">not present</i>',pb?esc(pb):'<i style="color:#bbb">not present</i>'];
    rows.push(`<div class="drow"><span class="slot">${ta}${ta!==tb?'<br>'+tb:''}</span><span>${ha}</span><span>${hb}</span></div>`);
   }
  }
 });
 document.getElementById('ptitle').textContent=`${va.id} vs ${vb.id} — ${diff} differing, ${same} identical paragraphs`;
 document.getElementById('pbody').innerHTML=rows.join('');
 const hs=document.getElementById('hidesame');
 const apply=()=>document.querySelectorAll('.hideable').forEach(e=>e.style.display=hs.checked?'none':'grid');
 hs.onchange=apply;apply();
 document.getElementById('overlay').style.display='block';
 document.getElementById('panel').style.display='flex';
};
document.getElementById('pclose').onclick=document.getElementById('overlay').onclick=()=>{
 document.getElementById('overlay').style.display='none';
 document.getElementById('panel').style.display='none';
};
</script></body></html>"""
open(os.path.join(OUT, "covenant_matrix.html"), "w", encoding="utf-8").write(
    tpl.replace("__DATA__", payload))
print(f"done {time.time()-t0:.0f}s")
