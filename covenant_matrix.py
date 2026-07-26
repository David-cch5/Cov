#!/usr/bin/env python3
"""
Covenant Version Matrix builder
--------------------------------
Scans a folder tree of PDFs (each named with a unique 4-digit CovID),
extracts text (OCR fallback for scanned files), clusters identical /
near-identical documents into Versions, and produces:

  Covenant_Matrix/covenant_matrix.html   - searchable visual for review
  Covenant_Matrix/covenant_matrix.json   - machine-readable (for AIs / tools)
  Covenant_Matrix/covenant_matrix.xlsx   - FileMaker import (Versions + CovID_Map)
  Covenant_Matrix/versions/V##.txt       - full text of each version
  Covenant_Matrix/issues.csv             - files that had problems

Usage:  python3 covenant_matrix.py <input_folder> [output_folder]
"""

import sys, os, re, json, hashlib, html, subprocess, tempfile, csv
from difflib import SequenceMatcher

import fitz  # PyMuPDF

CHUNK = 30000          # xlsx cell chunk size (limit is 32767)
OCR_THRESHOLD = 50     # chars/page below which we OCR
SIM_THRESHOLD = 0.995  # near-duplicate merge threshold (absorbs scan/OCR noise only;
                       # lower it if genuinely identical versions are being split)


# ---------------------------------------------------------------- extraction
def extract_text(path):
    """Returns (text, page_count, ocr_used)."""
    doc = fitz.open(path)
    pages = [p.get_text("text") for p in doc]
    text = "\n".join(pages)
    ocr = False
    if len(text.strip()) < OCR_THRESHOLD * max(len(doc), 1):
        # scanned - OCR each page
        ocr = True
        out = []
        with tempfile.TemporaryDirectory() as td:
            for i, page in enumerate(doc):
                pix = page.get_pixmap(dpi=200)
                png = os.path.join(td, f"p{i}.png")
                pix.save(png)
                r = subprocess.run(["tesseract", png, "stdout", "--psm", "1"],
                                   capture_output=True, text=True)
                out.append(r.stdout)
        text = "\n".join(out)
    n = len(doc)
    doc.close()
    return text, n, ocr


def normalize(text):
    """Aggressive normalization used only for comparing documents."""
    t = text.lower()
    t = re.sub(r"page \d+ of \d+", " ", t)
    t = re.sub(r"[^a-z0-9 ]+", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def clean_display(text):
    """Light cleanup of the raw text kept for display/storage."""
    t = text.replace("\r", "")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


# ---------------------------------------------------------------- clustering
class UF:
    def __init__(self, n): self.p = list(range(n))
    def find(self, x):
        while self.p[x] != x:
            self.p[x] = self.p[self.p[x]]; x = self.p[x]
        return x
    def union(self, a, b): self.p[self.find(a)] = self.find(b)


def cluster(norm_texts):
    """Group exact hashes, then merge near-identical groups. Returns list of member-index lists."""
    hashes = {}
    for i, t in enumerate(norm_texts):
        h = hashlib.md5(t.encode()).hexdigest()
        hashes.setdefault(h, []).append(i)
    reps = [members[0] for members in hashes.values()]
    uf = UF(len(reps))
    for a in range(len(reps)):
        ta = norm_texts[reps[a]][:5000]
        for b in range(a + 1, len(reps)):
            tb = norm_texts[reps[b]][:5000]
            la, lb = len(ta), len(tb)
            if min(la, lb) == 0 or min(la, lb) / max(la, lb) < 0.9:
                continue
            sm = SequenceMatcher(None, ta, tb)
            if sm.quick_ratio() >= SIM_THRESHOLD and sm.ratio() >= SIM_THRESHOLD:
                uf.union(a, b)
    merged = {}
    for gi, members in enumerate(hashes.values()):
        merged.setdefault(uf.find(gi), []).extend(members)
    return list(merged.values())


# ---------------------------------------------------------------- outputs
def build_xlsx(versions, mapping, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    bold = Font(name="Arial", bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    arial = Font(name="Arial")

    max_chunks = max((len(v["chunks"]) for v in versions), default=1)
    ws = wb.active; ws.title = "Versions"
    head = ["VersionID", "DocCount", "PageCount", "CharCount", "SampleCovID",
            "OCR_Used"] + [f"VersionText_{i+1}" for i in range(max_chunks)]
    ws.append(head)
    for v in versions:
        row = [v["version_id"], v["doc_count"], v["page_count"], v["char_count"],
               v["sample_covid"], "Yes" if v["ocr_used"] else "No"]
        row += v["chunks"] + [""] * (max_chunks - len(v["chunks"]))
        ws.append(row)
    for r in ws.iter_rows():
        for c in r: c.font = arial
    for c in ws[1]: c.font = bold; c.fill = fill
    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("CovID_Map")
    ws2.append(["CovID", "VersionID", "Filename", "RelativePath", "OCR_Used"])
    for m in mapping:
        ws2.append([m["covid"], m["version_id"], m["filename"],
                    m["relpath"], "Yes" if m["ocr"] else "No"])
    for r in ws2.iter_rows():
        for c in r: c.font = arial
    for c in ws2[1]: c.font = bold; c.fill = fill
    for col, w in zip("ABCDE", (10, 10, 30, 50, 10)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"
    wb.save(out_path)


def build_html(versions, mapping, stats, out_path):
    data = {"stats": stats,
            "versions": [{"id": v["version_id"], "docCount": v["doc_count"],
                          "pages": v["page_count"], "ocr": v["ocr_used"],
                          "covids": v["covids"], "paragraphs": v["paragraphs"]}
                         for v in versions]}
    payload = json.dumps(data).replace("</", "<\\/")
    tpl = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Covenant Version Matrix</title>
<style>
 :root{--blue:#1F4E79;--lt:#eef3f8;--hl:#ffe066}
 *{box-sizing:border-box}
 body{font-family:-apple-system,Segoe UI,Arial,sans-serif;margin:0;background:#f5f6f8;color:#222}
 header{background:var(--blue);color:#fff;padding:18px 28px}
 header h1{margin:0 0 4px;font-size:22px}
 .stats{display:flex;gap:26px;font-size:13px;opacity:.9}
 .stats b{font-size:17px;display:block}
 .bar{position:sticky;top:0;background:#fff;padding:12px 28px;box-shadow:0 1px 4px rgba(0,0,0,.12);z-index:5;display:flex;gap:12px;align-items:center}
 #q{flex:1;max-width:560px;padding:9px 14px;font-size:15px;border:1.5px solid #bbb;border-radius:8px}
 #hits{font-size:13px;color:#555}
 main{padding:20px 28px;max-width:1200px;margin:0 auto}
 .ver{background:#fff;border:1px solid #dde3ea;border-radius:10px;margin-bottom:16px;overflow:hidden}
 .vhead{display:flex;align-items:center;gap:14px;padding:13px 18px;cursor:pointer;background:var(--lt)}
 .vhead:hover{background:#e2ebf5}
 .vid{font-weight:700;color:var(--blue);font-size:17px;min-width:56px}
 .meta{font-size:12.5px;color:#555}
 .chips{padding:10px 18px;border-top:1px solid #e8edf2;display:flex;flex-wrap:wrap;gap:5px;max-height:132px;overflow:auto}
 .chip{background:var(--lt);border:1px solid #c9d7e6;border-radius:12px;padding:2px 9px;font-size:12px;font-family:ui-monospace,monospace}
 .chip.hit{background:var(--hl);border-color:#d4b106}
 .body{display:none;border-top:1px solid #e8edf2;padding:14px 18px;max-height:560px;overflow:auto}
 .ver.open .body{display:block}
 .para{display:flex;gap:12px;padding:5px 0;font-size:13.5px;line-height:1.55}
 .pn{color:#98a4b3;font-family:ui-monospace,monospace;font-size:11.5px;min-width:38px;text-align:right;flex-shrink:0;padding-top:2px}
 mark{background:var(--hl);padding:0 1px}
 .arrow{margin-left:auto;transition:.2s;color:#888}
 .ver.open .arrow{transform:rotate(90deg)}
 .hidden{display:none}
 .note{font-size:12px;color:#888;padding:4px 18px 10px}
</style></head><body>
<header><h1>Covenant Version Matrix</h1>
<div class="stats">
 <span><b id=s1></b>documents</span><span><b id=s2></b>versions</span>
 <span><b id=s3></b>OCR'd files</span><span><b id=s4></b>generated</span>
</div></header>
<div class="bar"><input id="q" placeholder="Search text, paragraph, or CovID (e.g. 4417, easement, setback)…" autofocus>
<span id="hits"></span></div>
<main id="list"></main>
<script>
const DATA = __DATA__;
const esc = s=>s.replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));
document.getElementById('s1').textContent=DATA.stats.docs;
document.getElementById('s2').textContent=DATA.versions.length;
document.getElementById('s3').textContent=DATA.stats.ocr;
document.getElementById('s4').textContent=DATA.stats.date;
const list=document.getElementById('list');
DATA.versions.forEach(v=>{
 const d=document.createElement('div');d.className='ver';d.dataset.id=v.id;
 d.innerHTML=`<div class="vhead"><span class="vid">${v.id}</span>
  <span class="meta">${v.docCount} document${v.docCount!==1?'s':''} · ${v.pages} pages · ${v.paragraphs.length} paragraphs${v.ocr?' · OCR':''}</span>
  <span class="arrow">▶</span></div>
  <div class="chips">${v.covids.map(c=>`<span class="chip" data-c="${c}">${c}</span>`).join('')}</div>
  <div class="body">${v.paragraphs.map((p,i)=>`<div class="para"><span class="pn">¶${i+1}</span><span class="pt">${esc(p)}</span></div>`).join('')}</div>`;
 d.querySelector('.vhead').onclick=()=>d.classList.toggle('open');
 list.appendChild(d);
});
const q=document.getElementById('q'),hits=document.getElementById('hits');
let t;q.addEventListener('input',()=>{clearTimeout(t);t=setTimeout(search,200)});
function search(){
 const s=q.value.trim().toLowerCase();
 document.querySelectorAll('mark').forEach(m=>m.replaceWith(m.textContent));
 document.querySelectorAll('.chip.hit').forEach(c=>c.classList.remove('hit'));
 let shown=0,pmatch=0;
 document.querySelectorAll('.ver').forEach(el=>{
  const v=DATA.versions.find(x=>x.id===el.dataset.id);
  if(!s){el.classList.remove('hidden','open');shown++;el.querySelectorAll('.para').forEach(p=>p.style.display='');return;}
  const cov=v.covids.filter(c=>c.toLowerCase().includes(s));
  let tp=0;
  el.querySelectorAll('.para').forEach(p=>{
   const el2=p.querySelector('.pt'),raw=el2.textContent;
   if(raw.toLowerCase().includes(s)){tp++;pmatch++;p.style.display='';
    el2.innerHTML=esc(raw).replace(new RegExp('('+s.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+')','gi'),'<mark>$1</mark>');
   } else p.style.display=s?'none':'';
  });
  cov.forEach(c=>el.querySelector(`.chip[data-c="${c}"]`)?.classList.add('hit'));
  if(tp||cov.length){el.classList.remove('hidden');el.classList.toggle('open',tp>0);shown++;}
  else el.classList.add('hidden');
 });
 hits.textContent=s?`${shown} version${shown!==1?'s':''}, ${pmatch} matching paragraph${pmatch!==1?'s':''}`:'';
}
</script></body></html>"""
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(tpl.replace("__DATA__", payload))


# ---------------------------------------------------------------- main
def main():
    if len(sys.argv) < 2:
        sys.exit("usage: covenant_matrix.py <input_folder> [output_folder]")
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(os.path.abspath(src.rstrip("/"))), "Covenant_Matrix")
    os.makedirs(out, exist_ok=True)
    os.makedirs(os.path.join(out, "versions"), exist_ok=True)

    pdfs = []
    for root, _, files in os.walk(src):
        for f in sorted(files):
            if f.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, f))
    if not pdfs:
        sys.exit(f"no PDFs found under {src}")

    docs, issues = [], []
    for i, p in enumerate(pdfs, 1):
        name = os.path.basename(p)
        m = re.search(r"\d{4}", name)
        covid = m.group(0) if m else ""
        if not covid:
            issues.append((name, "no 4-digit CovID in filename"))
        try:
            text, pages, ocr = extract_text(p)
        except Exception as e:
            issues.append((name, f"extraction failed: {e}"))
            continue
        if not text.strip():
            issues.append((name, "no text extracted (even after OCR)"))
        docs.append({"covid": covid or name, "path": p, "filename": name,
                     "relpath": os.path.relpath(p, src), "text": clean_display(text),
                     "norm": normalize(text), "pages": pages, "ocr": ocr})
        if i % 50 == 0 or i == len(pdfs):
            print(f"  extracted {i}/{len(pdfs)}", flush=True)

    print("clustering…", flush=True)
    groups = cluster([d["norm"] for d in docs])
    groups.sort(key=len, reverse=True)

    versions, mapping = [], []
    for gi, members in enumerate(groups, 1):
        vid = f"V{gi:02d}"
        members = sorted(members, key=lambda ix: docs[ix]["covid"])
        rep = docs[members[0]]
        paragraphs = [p.strip() for p in re.split(r"\n\s*\n", rep["text"]) if p.strip()]
        chunks = [rep["text"][i:i+CHUNK] for i in range(0, len(rep["text"]), CHUNK)] or [""]
        versions.append({"version_id": vid, "doc_count": len(members),
                         "page_count": rep["pages"], "char_count": len(rep["text"]),
                         "sample_covid": rep["covid"], "ocr_used": rep["ocr"],
                         "covids": [docs[m]["covid"] for m in members],
                         "paragraphs": paragraphs, "chunks": chunks,
                         "text": rep["text"]})
        for m in members:
            d = docs[m]
            mapping.append({"covid": d["covid"], "version_id": vid,
                            "filename": d["filename"], "relpath": d["relpath"],
                            "ocr": d["ocr"]})
        with open(os.path.join(out, "versions", f"{vid}.txt"), "w", encoding="utf-8") as f:
            f.write(rep["text"])

    mapping.sort(key=lambda m: m["covid"])
    import datetime
    stats = {"docs": len(docs), "ocr": sum(1 for d in docs if d["ocr"]),
             "date": datetime.date.today().isoformat()}

    build_xlsx(versions, mapping, os.path.join(out, "covenant_matrix.xlsx"))
    build_html(versions, mapping, stats, os.path.join(out, "covenant_matrix.html"))
    with open(os.path.join(out, "covenant_matrix.json"), "w", encoding="utf-8") as f:
        json.dump({"stats": stats,
                   "versions": [{k: v[k] for k in
                                 ("version_id", "doc_count", "page_count", "char_count",
                                  "sample_covid", "ocr_used", "covids", "paragraphs")}
                                for v in versions],
                   "covid_map": mapping}, f, indent=1)
    if issues:
        with open(os.path.join(out, "issues.csv"), "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f); w.writerow(["Filename", "Issue"]); w.writerows(issues)

    print(f"\nDone: {len(docs)} docs -> {len(versions)} versions")
    print(f"Issues: {len(issues)}")
    print(f"Output: {out}")


if __name__ == "__main__":
    main()
