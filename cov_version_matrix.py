#!/usr/bin/env python3
"""
Covenant Version Matrix — final builder.
Reads OCR'd text from _textcache/ (built by extract_parallel.py), clusters
documents into template versions (ignoring property-specific text), and writes:

  Covenant_Matrix/covenant_matrix.html   searchable visual
  Covenant_Matrix/covenant_matrix.json   machine-readable
  Covenant_Matrix/covenant_matrix.xlsx   FileMaker import (Versions + CovID_Map)
  Covenant_Matrix/versions/V###.txt      representative full text per version
  Covenant_Matrix/issues.csv             flags (near-empty OCR, duplicate CovIDs)

Usage: python3 cov_version_matrix.py [similarity_cut]   (default 0.87)
"""
import os, re, sys, json, glob, csv, datetime
from collections import Counter
import numpy as np
from scipy.sparse import csr_matrix
from scipy.cluster.hierarchy import linkage, fcluster
from scipy.spatial.distance import squareform

SRC = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SRC, "Covenant_Matrix")
CUT = float(sys.argv[1]) if len(sys.argv) > 1 else 0.87
K = 5          # shingle size (words)
MIN_DF = 5     # shingle must appear in >= MIN_DF docs to count as "template"

# ---------------------------------------------------------------- load
docs = []
for f in sorted(glob.glob(os.path.join(SRC, "_textcache", "*.json"))):
    docs.append(json.load(open(f, encoding="utf-8")))
n = len(docs)
print(f"{n} docs loaded")

def words(t):
    t = t.lower()
    t = re.sub(r"[0-9]+", "#", t)
    t = re.sub(r"[^a-z# ]+", " ", t)
    return [w for w in t.split() if len(w) > 1 or w == "#"]

df = Counter()
shsets = []
for d in docs:
    w = words(d["text"])
    s = {hash(" ".join(w[i:i+K])) for i in range(len(w)-K+1)}
    shsets.append(s)
    df.update(s)
common = {s for s, c in df.items() if c >= MIN_DF}
core = [s & common for s in shsets]

# ---------------------------------------------------------------- cluster
ids, rows, cols = {}, [], []
for i, s in enumerate(core):
    for sh in s:
        j = ids.setdefault(sh, len(ids))
        rows.append(i); cols.append(j)
M = csr_matrix((np.ones(len(rows), dtype=np.float32), (rows, cols)),
               shape=(n, len(ids)))
inter = (M @ M.T).toarray()
sz = np.array([len(s) for s in core], dtype=np.float32)
union = sz[:, None] + sz[None, :] - inter
J = np.where(union > 0, inter / np.maximum(union, 1), 0)
D = 1 - J; np.fill_diagonal(D, 0)
Z = linkage(squareform(D, checks=False), method="average")
lab = fcluster(Z, 1 - CUT, criterion="distance")
print(f"cut {CUT}: {len(set(lab))} versions")

clusters = {}
for i, l in enumerate(lab):
    clusters.setdefault(l, []).append(i)
ordered = sorted(clusters.values(), key=lambda m: (-len(m), min(docs[i]["covid"] for i in m)))

# ---------------------------------------------------------------- assemble
def clean(t):
    t = t.replace("\r", "")
    t = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", t)  # control chars (break xlsx)
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()

os.makedirs(OUT, exist_ok=True)
os.makedirs(os.path.join(OUT, "versions"), exist_ok=True)
CHUNK = 30000
versions, mapping, issues = [], [], []

covid_count = Counter(d["covid"] for d in docs)
for cov, c in sorted(covid_count.items()):
    if c > 1:
        issues.append((cov, f"{c} PDFs share this CovID"))

for gi, members in enumerate(ordered, 1):
    vid = f"V{gi:03d}"
    rep_i = max(members, key=lambda i: len(core[i]))
    rep = docs[rep_i]
    text = clean(rep["text"])
    # intra-cluster cohesion (min similarity of any member to rep)
    cohesion = float(min(J[rep_i][m] for m in members)) if len(members) > 1 else 1.0
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    chunks = [text[i:i+CHUNK] for i in range(0, len(text), CHUNK)] or [""]
    covids = sorted(docs[m]["covid"] for m in members)
    versions.append({"version_id": vid, "doc_count": len(members),
                     "page_count": rep["pages"], "char_count": len(text),
                     "sample_covid": rep["covid"], "cohesion": round(cohesion, 3),
                     "ocr_used": rep["ocr"], "covids": covids,
                     "paragraphs": paragraphs, "chunks": chunks})
    for m in members:
        d = docs[m]
        mapping.append({"covid": d["covid"], "version_id": vid,
                        "filename": d["filename"], "relpath": d["relpath"],
                        "pages": d["pages"], "ocr": d["ocr"],
                        "sim_to_rep": round(float(J[rep_i][m]), 3)})
        if len(d["text"].strip()) < 2000:
            issues.append((d["covid"], f"very little text extracted ({len(d['text'].strip())} chars, {d['pages']} pages) - likely plat map / handwriting / poor scan"))
        if d.get("error"):
            issues.append((d["covid"], "PDF could not be opened"))
    with open(os.path.join(OUT, "versions", f"{vid}.txt"), "w", encoding="utf-8") as f:
        f.write(text)

mapping.sort(key=lambda m: (m["covid"], m["filename"]))
stats = {"docs": n, "versions": len(versions),
         "multi": sum(1 for v in versions if v["doc_count"] > 1),
         "singletons": sum(1 for v in versions if v["doc_count"] == 1),
         "ocr": sum(1 for d in docs if d["ocr"]),
         "cut": CUT, "date": datetime.date.today().isoformat()}

# ---------------------------------------------------------------- xlsx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
wb = Workbook()
bold = Font(name="Arial", bold=True, color="FFFFFF")
fill = PatternFill("solid", fgColor="1F4E79")
arial = Font(name="Arial")
max_chunks = max(len(v["chunks"]) for v in versions)
ws = wb.active; ws.title = "Versions"
ws.append(["VersionID", "DocCount", "PageCount", "CharCount", "SampleCovID",
           "Cohesion", "OCR_Used"] + [f"VersionText_{i+1}" for i in range(max_chunks)])
for v in versions:
    ws.append([v["version_id"], v["doc_count"], v["page_count"], v["char_count"],
               v["sample_covid"], v["cohesion"], "Yes" if v["ocr_used"] else "No"]
              + v["chunks"] + [""] * (max_chunks - len(v["chunks"])))
for r in ws.iter_rows():
    for c in r: c.font = arial
for c in ws[1]: c.font = bold; c.fill = fill
for i in range(1, 8): ws.column_dimensions[get_column_letter(i)].width = 12
ws.freeze_panes = "A2"
ws2 = wb.create_sheet("CovID_Map")
ws2.append(["CovID", "VersionID", "Filename", "RelativePath", "Pages",
            "OCR_Used", "SimilarityToVersion"])
for m in mapping:
    ws2.append([m["covid"], m["version_id"], m["filename"], m["relpath"],
                m["pages"], "Yes" if m["ocr"] else "No", m["sim_to_rep"]])
for r in ws2.iter_rows():
    for c in r: c.font = arial
for c in ws2[1]: c.font = bold; c.fill = fill
for col, w in zip("ABCDEFG", (10, 10, 30, 40, 8, 10, 18)):
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A2"
wb.save(os.path.join(OUT, "covenant_matrix.xlsx"))

# ---------------------------------------------------------------- json
json.dump({"stats": stats,
           "versions": [{k: v[k] for k in ("version_id", "doc_count", "page_count",
                        "char_count", "sample_covid", "cohesion", "ocr_used",
                        "covids", "paragraphs")} for v in versions],
           "covid_map": mapping},
          open(os.path.join(OUT, "covenant_matrix.json"), "w", encoding="utf-8"),
          indent=1)

# ---------------------------------------------------------------- issues
with open(os.path.join(OUT, "issues.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(["CovID", "Issue"]); w.writerows(sorted(set(issues)))

# ---------------------------------------------------------------- html
data = {"stats": stats,
        "versions": [{"id": v["version_id"], "docCount": v["doc_count"],
                      "pages": v["page_count"], "ocr": v["ocr_used"],
                      "cohesion": v["cohesion"], "sample": v["sample_covid"],
                      "covids": v["covids"], "paragraphs": v["paragraphs"]}
                     for v in versions]}
payload = json.dumps(data).replace("</", "<\\/")
tpl = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Covenant Version Matrix</title>
<style>
 :root{--blue:#1F4E79;--lt:#eef3f8;--hl:#ffe066}
 *{box-sizing:border-box}
 body{font-family:-apple-system,Segoe UI,Arial,sans-serif;margin:0;background:#f5f6f8;color:#222}
 header{background:var(--blue);color:#fff;padding:16px 28px}
 header h1{margin:0 0 6px;font-size:21px}
 .stats{display:flex;gap:26px;font-size:12.5px;opacity:.92}
 .stats b{font-size:17px;display:block}
 .bar{position:sticky;top:0;background:#fff;padding:11px 28px;box-shadow:0 1px 4px rgba(0,0,0,.12);z-index:5;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
 #q{flex:1;min-width:260px;max-width:560px;padding:9px 14px;font-size:15px;border:1.5px solid #bbb;border-radius:8px}
 #hits{font-size:13px;color:#555}
 label.f{font-size:13px;color:#444;display:flex;gap:5px;align-items:center}
 main{padding:18px 28px;max-width:1250px;margin:0 auto}
 .ver{background:#fff;border:1px solid #dde3ea;border-radius:10px;margin-bottom:12px;overflow:hidden}
 .vhead{display:flex;align-items:center;gap:14px;padding:11px 16px;cursor:pointer;background:var(--lt)}
 .vhead:hover{background:#e2ebf5}
 .vid{font-weight:700;color:var(--blue);font-size:16px;min-width:56px}
 .meta{font-size:12.5px;color:#555}
 .chips{padding:8px 16px;border-top:1px solid #e8edf2;display:flex;flex-wrap:wrap;gap:4px;max-height:110px;overflow:auto}
 .chip{background:var(--lt);border:1px solid #c9d7e6;border-radius:12px;padding:1px 8px;font-size:11.5px;font-family:ui-monospace,monospace}
 .chip.hit{background:var(--hl);border-color:#d4b106}
 .body{display:none;border-top:1px solid #e8edf2;padding:12px 16px;max-height:540px;overflow:auto}
 .ver.open .body{display:block}
 .para{display:flex;gap:12px;padding:4px 0;font-size:13.5px;line-height:1.5}
 .pn{color:#98a4b3;font-family:ui-monospace,monospace;font-size:11px;min-width:40px;text-align:right;flex-shrink:0;padding-top:2px}
 mark{background:var(--hl);padding:0 1px}
 .arrow{margin-left:auto;transition:.2s;color:#888}
 .ver.open .arrow{transform:rotate(90deg)}
 .hidden{display:none}
 .more{font-size:12px;color:var(--blue);cursor:pointer;padding:6px 0}
</style></head><body>
<header><h1>Covenant Version Matrix</h1>
<div class="stats">
 <span><b id=s1></b>documents</span><span><b id=s2></b>versions</span>
 <span><b id=s3></b>shared versions</span><span><b id=s4></b>one-off documents</span>
 <span><b id=s5></b>generated</span>
</div></header>
<div class="bar"><input id="q" placeholder="Search text or CovID (e.g. 4417, transfer fee, gross sales price)…" autofocus>
<label class="f"><input type="checkbox" id="multi" checked> hide one-off versions</label>
<span id="hits"></span></div>
<main id="list"></main>
<script>
const DATA = __DATA__;
const esc=s=>s.replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));
s1.textContent=DATA.stats.docs;s2.textContent=DATA.stats.versions;
s3.textContent=DATA.stats.multi;s4.textContent=DATA.stats.singletons;
s5.textContent=DATA.stats.date;
const list=document.getElementById('list'),els={};
DATA.versions.forEach(v=>{
 const d=document.createElement('div');d.className='ver';d.dataset.id=v.id;
 if(v.docCount===1)d.dataset.single="1";
 d.innerHTML=`<div class="vhead"><span class="vid">${v.id}</span>
  <span class="meta">${v.docCount} doc${v.docCount!==1?'s':''} · ${v.pages} pages · ${v.paragraphs.length} ¶ · sample ${v.sample}${v.docCount>1?' · cohesion '+v.cohesion:''}</span>
  <span class="arrow">▶</span></div>
  <div class="chips">${v.covids.slice(0,400).map(c=>`<span class="chip" data-c="${c}">${c}</span>`).join('')}</div>
  <div class="body"></div>`;
 d.querySelector('.vhead').onclick=()=>{d.classList.toggle('open');if(d.classList.contains('open'))render(v,d,curQ());};
 els[v.id]=d;list.appendChild(d);
});
function curQ(){return document.getElementById('q').value.trim().toLowerCase()}
function render(v,el,s){
 const body=el.querySelector('.body');
 const rx=s?new RegExp('('+s.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+')','gi'):null;
 let out=[];
 v.paragraphs.forEach((p,i)=>{
  if(s&&!p.toLowerCase().includes(s))return;
  let h=esc(p);if(rx)h=h.replace(rx,'<mark>$1</mark>');
  out.push(`<div class="para"><span class="pn">¶${i+1}</span><span>${h}</span></div>`);
 });
 if(s&&!out.length)out.push('<div class="para"><span></span><span style="color:#999">CovID match only — no text match in this version</span></div>');
 body.innerHTML=out.join('');
}
const q=document.getElementById('q'),hits=document.getElementById('hits'),multi=document.getElementById('multi');
let t;q.addEventListener('input',()=>{clearTimeout(t);t=setTimeout(search,250)});
multi.addEventListener('change',search);
function search(){
 const s=curQ();let shown=0,pm=0;
 DATA.versions.forEach(v=>{
  const el=els[v.id];
  el.querySelectorAll('.chip.hit').forEach(c=>c.classList.remove('hit'));
  if(!s){
   const hide=multi.checked&&v.docCount===1;
   el.classList.toggle('hidden',hide);el.classList.remove('open');
   if(!hide)shown++;return;
  }
  const cov=v.covids.filter(c=>c.includes(s));
  let tp=0;v.paragraphs.forEach(p=>{if(p.toLowerCase().includes(s)){tp++;pm++;}});
  const match=tp>0||cov.length>0;
  const hide=!match||(multi.checked&&v.docCount===1&&!cov.length&&!tp);
  el.classList.toggle('hidden',hide);
  if(!hide){shown++;el.classList.toggle('open',true);render(v,el,s);
   cov.forEach(c=>el.querySelector(`.chip[data-c="${c}"]`)?.classList.add('hit'));}
 });
 hits.textContent=s?`${shown} versions, ${pm} matching paragraphs`:'';
}
search();
</script></body></html>"""
open(os.path.join(OUT, "covenant_matrix.html"), "w", encoding="utf-8").write(
    tpl.replace("__DATA__", payload))

print(f"versions: {stats['versions']} ({stats['multi']} shared, {stats['singletons']} one-off)")
print(f"issues: {len(set(issues))}")
print("output:", OUT)
