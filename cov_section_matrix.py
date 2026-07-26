#!/usr/bin/env python3
"""
Section-based Covenant Version Matrix.

Per user spec:
 - Only the NUMBERED SECTIONS after "WITNESSETH:" are compared (1..30/31...).
 - Grantor info (above WITNESSETH), signature/notary blocks and Exhibit A legal
   descriptions are EXCLUDED and never cause a new version.
 - Sections are aligned across documents into SLOTS (S001...) named by their
   typical section number; different language within a slot = VARIANTS (S012_v2).
 - A VERSION = a distinct combination of section variants. Two docs whose
   sections all carry the same language (even with scan noise) are one version.

Reads _textcache/*.json.  Resumable: run repeatedly until "ALL DONE".
Outputs into Covenant_Matrix/: covenant_matrix.xlsx, covenant_matrix.html,
covenant_matrix.json, versions/V###.txt, issues.csv
"""
import os, re, sys, json, glob, csv, time, pickle, datetime
import numpy as np
from collections import Counter, defaultdict
from difflib import SequenceMatcher

SRC = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(SRC, "Covenant_Matrix")
BUDGET = 36
VARIANT_SIM = 0.78          # difflib: same language (OCR noise) vs real rewording
SIG_MERGE = 0.85            # signature Jaccard to merge noise-split versions
t0 = time.time()

# ============================================================ 1. parse sections
WIT = re.compile(r"W\s*[I1L|]\s*T\s*N\s*E\s*S\s*S\s*E\s*T\s*H", re.I)
# fuzzy: OCR renders WHEREOF as WHEREOPF/WHEREOEF etc.
SIG = re.compile(r"IN\s+WITNESS\s+WHERE\w{0,4}|WITNESS\s+MY\s+HAND|SIGNED\s+AND\s+DELIVERED", re.I)
# Exhibit A / legal-description header. Uppercase form matches anywhere on a line;
# title-case form only when it is alone on its line (a real header, not an inline
# "described in Exhibit A" reference). Only honored in the latter part of a doc.
EXH = re.compile(r"(?m)^\W{0,8}(?:EXHIBIT|SCHEDULE)\s*[\"'“”‘’]{0,2}\s*A\b"
                 r"|^\W{0,8}LEGAL\s+DESCRIPTION"
                 r"|^\W{0,8}(?:Exhibit|Schedule)\s*[\"'“”‘’]{0,2}\s*[Aa][\"'“”‘’]{0,2}\s*\W{0,6}$")
SURVEY = re.compile(r"\bTHENCE\b|POINT OF BEGINNING|METES AND BOUNDS"
                    r"|BEING A (?:TRACT|PARCEL) OF LAND|FIELD NOTES?"
                    r"|Property ID\s*\d|,\s*Acres\s+[\d.]"
                    r"|according to the (?:plat|map)\b"
                    r"|recorded in (?:Volume|Book|Cabinet|Plat)", re.I)

# ---- OCR cleanup: frequent, unambiguous misreads mined from this corpus -------
OCR_FIXES = [
    (re.compile(r"(?<=[A-Za-z])[!|](?=[A-Za-z])"), "l"),          # shal!l, Dec!arant
    (re.compile(r"(?<=[a-z])1(?=[a-z])"), "l"),                    # sha1l
    (re.compile(r"\bshal[il!f1|]{1,2}\b|\bshail\b|\bshull\b|\bshalt\b|\bsball\b|\bshal\b"), "shall"),
    (re.compile(r"\bshall[!|]+(?=[\s,.;)]|$)"), "shall"),
    (re.compile(r"\bShal[il!f1|]{1,2}\b|\bShail\b|\bSball\b"), "Shall"),
    (re.compile(r"\bhercin\b|\bberein\b"), "herein"),
    (re.compile(r"\bthercof\b"), "thereof"),
    (re.compile(r"\bbereto\b|\bhercto\b"), "hereto"),
    (re.compile(r"\bbereof\b|\bhercof\b"), "hereof"),
    (re.compile(r"\bTrust[ce][ce]\b|\bTrusiee\b|\bTrusice\b|\b[lI]rustee\b"), "Trustee"),
    (re.compile(r"\bDec[il][ae]rant\b|\bDectarant\b|\bDecla[fy]ant\b"), "Declarant"),
    (re.compile(r"\bBencficiary\b|\bBeneticiary\b|\bBenefictary\b"), "Beneficiary"),
    (re.compile(r"\bBeneficiar[il][ce]s\b|\bBencficiar(?:ie|ic)s\b|\bBeneficianes\b|\bBene[lt]iciaries\b"), "Beneficiaries"),
    (re.compile(r"\bPropes?ty\b|\bProperfy\b|\bProperiy\b"), "Property"),
    (re.compile(r"\b([Cc])onvevan[ce]e\b|\b([Cc])onveyanee\b"), lambda m: (m.group(1) or m.group(2)) + "onveyance"),
    (re.compile(r"\bReconve[vy]an[ce]e\b|\bRecon[vy]e[vy]ance\b"), "Reconveyance"),
    (re.compile(r"\b([Ii])nstrament\b|\bInsirument\b|\blnstrument\b"), "Instrument"),
    (re.compile(r"\bDec[til][ao]ration\b"), "Declaration"),
    (re.compile(r"\b([Aa])t{1,2}[io]mey"), lambda m: m.group(1) + "ttorney"),
    (re.compile(r"\bQwner\b|\bOwncr\b"), "Owner"),
    (re.compile(r"\bWHEREO[PE]F\b"), "WHEREOF"),
    (re.compile(r"\bDERINITIONS\b|\bDEFINITIGNS\b"), "DEFINITIONS"),
    (re.compile(r"\bwansfer\b"), "transfer"),
    (re.compile(r"\b[EF]stoppe[lit]?\b"), "Estoppel"),
    (re.compile(r"\bforec[ti]osure\b"), "foreclosure"),
    (re.compile(r"\bintcrest\b|\bnterest\b|\binteres\b"), "interest"),
    (re.compile(r"\b[il]censor\b"), "Licensor"),
    (re.compile(r"\ba([A-Z][a-z]+)\b"), r"a \1"),               # aClosing -> a Closing
]
def ocr_fix(t):
    for rx, rep in OCR_FIXES:
        t = rx.sub(rep, t)
    return t

# page footers / recorder stamps that OCR leaves inside the text stream.
# Removed line-wise BEFORE section parsing so they never join a section.
FOOTERS = [
    # any line that STARTS like a File-number stamp goes entirely (recording
    # numbers, timestamps and clerk names often trail it)
    re.compile(r"(?mi)^\W{0,4}[a-z]?\s{0,2}fil[a-z]{1,4}\s*[#®=?t\]_]*\s*\d.*$"),
    re.compile(r"(?mi)^\s*Deed\s+Book\s+\d+\s+Page\s+\d+\s*$"),
    re.compile(r"(?mi)^.{0,40}Doc\.?\s*ID:?\s*[\d\-]+\s*$"),
    re.compile(r"(?mi)^\W{0,4}file\s*[#®=?]\s*\d*\s*_{0,3}\s*$"),
    re.compile(r"(?m)^\W{0,3}#\d{5,}[, ]+\d+\s+OF\s+\d+\s*$"),
    re.compile(r"(?mi)^\s*page\s+\d+(\s+of\s+\d+)?\s*$"),
    re.compile(r"(?m)^\s*\d{1,3}\s*$"),                          # bare page numbers
    re.compile(r"(?m)^\s*\d+\s+OPR\s+[\d ]+$"),
    re.compile(r"(?mi)^\s*BOOK:?\s*\d+\s+PAGE:?\s*\d+\s*$"),
    re.compile(r"(?mi)^\s*[A-Z]{3},?\s+\d{1,2},?\s+\d{4}\s+\d{1,2}:\d{2}\s*[AP]M.*$"),
]
def scrub_footers(t):
    for rx in FOOTERS:
        t = rx.sub("", t)
    return t
# NOTE: "(" is excluded from the junk prefix so inline enumerations like
# "(1) assessed... (2) computed..." can never be mistaken for section starts.
SECRE = re.compile(
    r"(?m)^[ \t]{0,8}[^\w\n(]{0,4}?(?:[a-z]{1,2}[ \t]+)?[^\w\n(]{0,3}"
    r"(\d{1,2}|[IiLl!|])\s*[\.\),:]\s+(?=\S)")
MAX_SECNO = 39                     # covenants number to ~31; beyond = riders/exhibits

def tonum(s): return 1 if s in "IiLl!|" else int(s)

def parse_sections(t):
    m = WIT.search(t)
    start = m.end() if m else 0
    if not m:
        m2 = re.search(r"NOW,?\s+THEREFORE", t, re.I)
        start = m2.start() if m2 else 0
    body = t[start:]
    s = SIG.search(body)
    if s: body = body[:s.start()]
    # cut Exhibit A / Schedule A / legal description in the latter part of the doc
    for m in EXH.finditer(body):
        if m.start() > 0.35 * len(body):
            body = body[:m.start()]
            break
    cands = [(x.start(), tonum(x.group(1))) for x in SECRE.finditer(body)]
    cands = [c for c in cands if c[1] <= MAX_SECNO]
    best, best_score = [], -1e9
    blen = max(len(body), 1)
    for i0 in range(len(cands)):
        if cands[i0][1] != 1: continue
        run = [cands[i0]]
        for k in range(i0 + 1, len(cands)):
            if 1 <= cands[k][1] - run[-1][1] <= 4:
                run.append(cands[k])
        consec = sum(1 for a, b in zip(run, run[1:]) if b[1] - a[1] == 1)
        gaps = sum(b[1] - a[1] - 1 for a, b in zip(run, run[1:]))
        # prefer long, densely-numbered runs that start near the body's beginning
        score = 2 * len(run) + 3 * consec - gaps - 12 * (run[0][0] / blen)
        if score > best_score:
            best, best_score = run, score
    secs = [(no, body[pos: best[i+1][0] if i+1 < len(best) else len(body)].strip())
            for i, (pos, no) in enumerate(best)]
    # unlabeled legal descriptions can trail the final section: trim at the first
    # strong surveying marker (metes-and-bounds language never appears in sections)
    if secs:
        m = SURVEY.search(secs[-1][1], 300)
        if m:
            secs[-1] = (secs[-1][0], secs[-1][1][:m.start()].strip())
    start_off = best[0][0] if best else len(body)
    return secs, body.strip(), start_off

def norm(t):
    t = t.lower()
    t = re.sub(r"[0-9]+", "#", t)
    t = re.sub(r"[^a-z# ]+", " ", t)
    # drop 1-char tokens: mostly scan-margin junk ("| i t e"), heavy OCR noise
    return " ".join(w for w in t.split() if len(w) > 1)

CACHE_DIR = os.environ.get("COV_CACHE", os.path.join(SRC, "_textcache"))

# ---- auto-mined OCR corrections: frequent tokens that are 1-2 edits away from a
# much more frequent corpus word get corrected to it (data-driven, cached for
# inspection in _ocr2_work/autofix.json)
def build_autofix(files):
    af_path = os.path.join(SRC, "_ocr2_work", "autofix_v4.json")
    if os.path.exists(af_path):
        return json.load(open(af_path))
    freq = Counter()
    for f in files:
        d = json.load(open(f, encoding="utf-8"))
        freq.update(w.lower() for w in re.findall(r"[A-Za-z]{4,}", d["text"]))
    try:
        from english_words import get_english_words_set
        ENGLISH = get_english_words_set(["web2"], lower=True)
    except ImportError:
        ENGLISH = set()
    def stems(w):
        out = {w}
        for suf in ("ing", "ed", "es", "s", "d", "ly"):
            if w.endswith(suf) and len(w) - len(suf) >= 4:
                out.add(w[:-len(suf)])
        return out
    common = {w for w, c in freq.items() if c >= 50}
    by_shape = defaultdict(list)
    for w in common:
        by_shape[(w[0], len(w))].append(w)
    fixes = {}
    for t, c in freq.items():
        if c < 10 or c > 500 or t in common or len(t) < 6:
            continue
        if t in ENGLISH:                       # real word: never "correct" it
            continue
        best, br = None, 0.90
        for L in (len(t) - 1, len(t), len(t) + 1):
            for w in by_shape.get((t[0], L), []):
                if freq[w] < 30 * c:
                    continue
                if stems(t) & stems(w):        # plural/inflection, not a misread
                    continue
                r = SequenceMatcher(None, t, w).ratio()
                if r > br:
                    best, br = w, r
        if best:
            fixes[t] = best
    os.makedirs(os.path.join(SRC, "_ocr2_work"), exist_ok=True)
    json.dump(fixes, open(af_path, "w"), indent=1)
    return fixes

PARSE_CKPT = "/tmp/sec_parse.pkl"
if os.path.exists(PARSE_CKPT):
    docs = pickle.load(open(PARSE_CKPT, "rb"))
else:
    cache_files = sorted(glob.glob(os.path.join(CACHE_DIR, "*.json")))
    AUTOFIX = build_autofix(cache_files)
    print(f"auto-mined OCR corrections: {len(AUTOFIX)}")
    AF_RX = re.compile(r"\b[A-Za-z]{4,}\b")
    def autofix(t):
        def rep(m):
            w = m.group(0); lw = w.lower()
            fx = AUTOFIX.get(lw)
            if not fx: return w
            if w.isupper(): return fx.upper()
            if w[0].isupper(): return fx.capitalize()
            return fx
        return AF_RX.sub(rep, t)
    docs = []
    for f in cache_files:
        d = json.load(open(f, encoding="utf-8"))
        secs, body, start_off = parse_sections(scrub_footers(autofix(ocr_fix(d["text"]))))
        docs.append({"covid": d["covid"], "filename": d["filename"],
                     "relpath": d["relpath"], "pages": d["pages"], "ocr": d["ocr"],
                     "secs": secs, "body": body, "start_off": start_off,
                     "parsed": len(secs) >= 5})
    pickle.dump(docs, open(PARSE_CKPT, "wb"))
n_parsed = sum(d["parsed"] for d in docs)
print(f"{len(docs)} docs, {n_parsed} parsed into sections  {time.time()-t0:.0f}s")

# all sections from parsed docs
sections, sowner = [], []          # text ; (doc_idx, position_in_doc, section_no)
for di, d in enumerate(docs):
    if not d["parsed"]: continue
    for pi, (no, txt) in enumerate(d["secs"]):
        sections.append(txt)
        sowner.append((di, pi, no))
snorms = [norm(s) for s in sections]
# fill-in mask — used ONLY to decide whether two variants share the same template
# language. Drops words adjacent to numbers (addresses, dates, percentages) and
# words preceding entity suffixes (beneficiary/company names), which vary
# per document without being a language change.
ENTITY_SUF = {"llc", "ltd", "inc", "corp", "co", "company", "partnership",
              "lp", "llp", "trust"}
def masked(i):
    # like norm() but keeps single letters so enumeration markers (a. b. c.) survive
    t = sections[i].lower()
    t = re.sub(r"[0-9]+", "#", t)
    t = re.sub(r"[^a-z# ]+", " ", t)
    ws = t.split()
    drop = [False] * len(ws)
    # whole enumerated entries that contain a number or entity suffix are
    # fill-in list items (beneficiary name + address + percentage) — mask entirely.
    # Long entries (>25 words) are prose sub-clauses and are kept.
    marks = [j for j, w in enumerate(ws) if len(w) == 1 and w != "#"]
    for mi, j in enumerate(marks):
        end = marks[mi + 1] if mi + 1 < len(marks) else len(ws)
        span = ws[j:end]
        if len(span) <= 25 and any("#" in w or w in ENTITY_SUF for w in span):
            for k in range(j, end):
                drop[k] = True
    for j, w in enumerate(ws):
        if "#" in w:
            for k in range(max(0, j - 2), min(len(ws), j + 3)):
                drop[k] = True
        if w in ENTITY_SUF:
            for k in range(max(0, j - 4), j + 1):
                drop[k] = True
    # gap-closing: a word squeezed between masked spans (within 2 on both sides)
    # is list residue (city names, stray person names), not prose — mask it too
    for _ in range(3):
        changed = False
        for j in range(len(ws)):
            if drop[j]:
                continue
            left = any(drop[k] for k in range(max(0, j - 2), j))
            right = any(drop[k] for k in range(j + 1, min(len(ws), j + 3)))
            if left and right:
                drop[j] = True
                changed = True
        if not changed:
            break
    return " ".join(w for j, w in enumerate(ws) if not drop[j] and len(w) > 1)
print(f"{len(sections)} sections")

# ============================================================ 2. slot clustering
SLOT_CKPT = "/tmp/sec_slots.pkl"
if os.path.exists(SLOT_CKPT):
    X, slot_lab = pickle.load(open(SLOT_CKPT, "rb"))
    print(f"slot checkpoint loaded  {time.time()-t0:.0f}s")
else:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.preprocessing import normalize as l2
    from scipy.sparse.csgraph import connected_components
    from scipy.cluster.hierarchy import linkage, fcluster
    from scipy.spatial.distance import squareform
    from scipy.sparse import csr_matrix, vstack as spvstack
    XCKPT = "/tmp/sec_X.pkl"
    if os.path.exists(XCKPT):
        X = pickle.load(open(XCKPT, "rb"))
    else:
        X = l2(TfidfVectorizer(analyzer="char_wb", ngram_range=(4, 4),
                               max_features=200000, sublinear_tf=True).fit_transform(snorms))
        pickle.dump(X, open(XCKPT, "wb"))
    print(f"tfidf  {time.time()-t0:.0f}s")
    # chunked, resumable neighbor graph (cosine > 0.60)
    GCKPT = "/tmp/sec_G.pkl"
    CH = 1000
    if os.path.exists(GCKPT):
        rows_done, blocks = pickle.load(open(GCKPT, "rb"))
    else:
        rows_done, blocks = 0, []
    N = X.shape[0]
    while rows_done < N:
        if time.time() - t0 > BUDGET - 8:
            pickle.dump((rows_done, blocks), open(GCKPT + ".tmp", "wb"))
            os.replace(GCKPT + ".tmp", GCKPT)
            print(f"PARTIAL graph {rows_done}/{N} — rerun to continue")
            sys.exit(1)
        hi = min(rows_done + CH, N)
        B = (X[rows_done:hi] @ X.T)
        B.data[B.data <= 0.60] = 0
        B.eliminate_zeros()
        blocks.append(B)
        rows_done = hi
        print(f"  graph {rows_done}/{N}  {time.time()-t0:.0f}s", flush=True)
    G = spvstack(blocks).tocsr()
    ncomp, comp = connected_components(G > 0, directed=False)
    print(f"{ncomp} components  {time.time()-t0:.0f}s")

    next_lab = 0
    slot_lab = np.empty(N, dtype=np.int64)
    def assign(members, level=0):
        """avg-linkage cut at sim 0.60; oversized components pre-split at rising thresholds"""
        global next_lab
        if len(members) <= 2:
            for i in members: slot_lab[i] = next_lab
            next_lab += 1
            return
        if len(members) > 3000:
            ths = [0.70, 0.80, 0.90]
            if level < len(ths):
                sub = G[members][:, members]
                nc, cl = connected_components(sub > ths[level], directed=False)
                if nc > 1:
                    gs = defaultdict(list)
                    for m, c in zip(members, cl): gs[c].append(m)
                    for g in gs.values(): assign(g, level + 1)
                    return
                assign(members, level + 1)
                return
            for i in members: slot_lab[i] = next_lab
            next_lab += 1
            return
        Xs = X[members]
        Dm = np.clip(1 - (Xs @ Xs.T).toarray(), 0, None)
        np.fill_diagonal(Dm, 0)
        sub = fcluster(linkage(squareform(Dm, checks=False), method="average"),
                       0.40, criterion="distance")
        for i, s in zip(members, sub): slot_lab[i] = next_lab + s - 1
        next_lab += int(sub.max())
    comp_members = defaultdict(list)
    for i, l in enumerate(comp): comp_members[l].append(i)
    for members in comp_members.values():
        assign(members)
    pickle.dump((X, slot_lab), open(SLOT_CKPT, "wb"))
    print(f"{len(set(slot_lab))} slots  {time.time()-t0:.0f}s")

slot_members = defaultdict(list)
for i, l in enumerate(slot_lab): slot_members[l].append(i)
# slot naming by modal section number, then usage
def modal_no(members): return Counter(sowner[i][2] for i in members).most_common(1)[0][0]
slot_order = sorted(slot_members, key=lambda l: (modal_no(slot_members[l]),
                                                 -len(slot_members[l])))
slot_id = {l: f"S{r+1:03d}" for r, l in enumerate(slot_order)}
slot_no = {slot_id[l]: modal_no(slot_members[l]) for l in slot_members}

# ============================================================ 3. variants
VAR_CKPT = "/tmp/sec_var.pkl"
if os.path.exists(VAR_CKPT):
    variant_of, variants, done_slots = pickle.load(open(VAR_CKPT, "rb"))
    print(f"variant checkpoint: {len(done_slots)} slots done")
else:
    variant_of, variants, done_slots = {}, {}, set()
for l in sorted(slot_members):
    if l in done_slots: continue
    if time.time() - t0 > BUDGET:
        pickle.dump((variant_of, variants, done_slots), open(VAR_CKPT, "wb"))
        print(f"PARTIAL variants {len(done_slots)}/{len(slot_members)} — rerun to continue")
        sys.exit(1)
    members = sorted(slot_members[l], key=lambda i: -len(snorms[i]))
    sid = slot_id[l]
    if len(members) == 1:
        variants[(sid, 1)] = {"rep": members[0], "members": [members[0]]}
        variant_of[members[0]] = (sid, 1)
        done_slots.add(l)
        continue
    Xs = X[members]
    S = (Xs @ Xs.T).toarray()
    big = len(members) > 400
    lvo, lvar, leaders = {}, {}, []
    for li, i in enumerate(members):
        vn = None
        if leaders:
            order = np.argsort(-S[li, leaders])[:40]
            for oi in order:
                c = S[li, leaders[oi]]
                if c >= 0.95 or (big and c >= 0.80): vn = oi + 1; break
                if c < 0.55 or big: break
                a, b = snorms[i][:1500], snorms[members[leaders[oi]]][:1500]
                sm = SequenceMatcher(None, a, b)
                if sm.quick_ratio() >= VARIANT_SIM and sm.ratio() >= VARIANT_SIM:
                    vn = oi + 1; break
        if vn is None:
            leaders.append(li); vn = len(leaders)
            lvar[(sid, vn)] = {"rep": i, "members": []}
        lvar[(sid, vn)]["members"].append(i)
        lvo[i] = (sid, vn)
    variants.update(lvar); variant_of.update(lvo)
    done_slots.add(l)
print(f"variants: {len(variants)}  {time.time()-t0:.0f}s")

# renumber variants by usage
vcount = {k: len({sowner[i][0] for i in d["members"]}) for k, d in variants.items()}
new_no = {}
for sid in {s for s, _ in variants}:
    for newn, k in enumerate(sorted([k for k in variants if k[0] == sid],
                                    key=lambda k: (-vcount[k], k[1])), 1):
        new_no[k] = newn
variant_of = {i: (s, new_no[(s, v)]) for i, (s, v) in variant_of.items()}
variants = {(s, new_no[(s, v)]): d for (s, v), d in variants.items()}

# ============================================================ 3b. topics from headings
# every section starts "12. TRUSTEE RIGHTS. ..." -> the heading is the stable,
# human-meaningful key; section NUMBERS shift between versions.
def extract_heading(t):
    m = re.match(r"^\W{0,10}(?:\d{1,2}|[IiLl!|])\s*[\.\),:]\s*(.{3,80}?)[\.\n:]", t, re.S)
    if not m: return None
    h = m.group(1)
    letters = re.sub(r"[^A-Za-z]", "", h)
    if len(letters) < 3: return None
    if sum(c.isupper() for c in letters) / len(letters) < 0.6: return None
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z ]", " ", h.upper())).strip()

sec_head = [extract_heading(s) for s in sections]
slot_head = {}
for l, members in slot_members.items():
    c = Counter(h for h in (sec_head[i] for i in members) if h)
    slot_head[slot_id[l]] = c.most_common(1)[0][0] if c else None

head_w = Counter()
for l, members in slot_members.items():
    h = slot_head[slot_id[l]]
    if h: head_w[h] += len(members)
def head_key(h):
    """normalize terminology shifts between template generations for merging"""
    return h.replace("RECONVEYANCE", "TRANSFER")
topic_leaders, topic_of_head = [], {}
for h, _ in head_w.most_common():
    hit = None
    hk = head_key(h)
    for tl in topic_leaders:
        tk = head_key(tl)
        r = SequenceMatcher(None, hk, tk).ratio()
        # 0.87: high enough that BENEFICIARIES vs BENEFICIARY DUTIES stay apart,
        # low enough that OCR-mangled headings still merge. Word-boundary prefix
        # handles renamed topics (LIEN AND PRIORITY -> LIEN AND PRIORITY, LIABILITY...)
        prefix = (min(len(hk), len(tk)) >= 8
                  and (hk.startswith(tk + " ") or tk.startswith(hk + " ")
                       or hk == tk))
        if r >= 0.87 or prefix:
            hit = tl; break
    if hit is None:
        topic_leaders.append(h); hit = h
    topic_of_head[h] = hit
topic_of_slot = {}
for sid, h in slot_head.items():
    topic_of_slot[sid] = topic_of_head[h] if h else f"(UNNAMED {sid})"
# typical section number per topic (weighted modal)
topic_no_votes = defaultdict(Counter)
for l, members in slot_members.items():
    tp = topic_of_slot[slot_id[l]]
    for i in members:
        topic_no_votes[tp][sowner[i][2]] += 1
topic_no = {tp: c.most_common(1)[0][0] for tp, c in topic_no_votes.items()}
print(f"{len(topic_leaders)} named topics  {time.time()-t0:.0f}s")

# ============================================================ 4. versions from signatures
doc_sig = {}                       # doc idx -> frozenset of "S###_v#"
doc_secmap = defaultdict(list)     # doc idx -> [(secno, sid, vn, sec_idx)]
for i, (di, pi, no) in enumerate(sowner):
    sid, vn = variant_of[i]
    doc_secmap[di].append((no, sid, vn, i))
for di in doc_secmap:
    doc_sig[di] = frozenset(f"{s}_v{v}" for _, s, v, _ in doc_secmap[di])

sig_groups = defaultdict(list)
for di, sig in doc_sig.items(): sig_groups[sig].append(di)
sigs = sorted(sig_groups, key=lambda s: (-len(sig_groups[s]), sorted(s)[:3]))
# merge signatures that differ only by residual noise (Jaccard >= SIG_MERGE)
parent = list(range(len(sigs)))
def find(x):
    while parent[x] != x: parent[x] = parent[parent[x]]; x = parent[x]
    return x
for a in range(len(sigs)):
    for b in range(a + 1, len(sigs)):
        sa, sb = sigs[a], sigs[b]
        inter = len(sa & sb)
        if inter and inter / len(sa | sb) >= SIG_MERGE:
            parent[find(b)] = find(a)
merged = defaultdict(list)
for gi in range(len(sigs)): merged[find(gi)].extend(sig_groups[sigs[gi]])
frags = sorted(merged.values(), key=lambda m: (-len(m), min(docs[d]["covid"] for d in m)))
print(f"{len(frags)} signature fragments  {time.time()-t0:.0f}s")

# ---- consolidate fragments into true versions -------------------------------
# Bad scans split off tiny fragments. Fragments are clustered on their
# majority-voted section text; clusters with >= MIN_CORE_DOCS docs become CORE
# versions, every other fragment attaches to its nearest core (confidence
# recorded, ATTACH_MIN floor keeps truly alien documents separate).
CORE_CUT = 0.85
MIN_CORE_DOCS = 3
ATTACH_MIN = 0.50

def frag_votes(members):
    votes = defaultdict(Counter)
    for di in members:
        for _, sid, vn, _ in doc_secmap[di]:
            votes[sid][vn] += 1
    return votes

frag_body = []
for members in frags:
    votes = frag_votes(members)
    parts = []
    for sid, cnt in votes.items():
        vn = cnt.most_common(1)[0][0]
        parts.append((slot_no[sid], snorms[variants[(sid, vn)]["rep"]]))
    frag_body.append(" ".join(p for _, p in sorted(parts, key=lambda x: x[0])))

from sklearn.feature_extraction.text import TfidfVectorizer as _TV
from sklearn.preprocessing import normalize as _l2n
from scipy.cluster.hierarchy import linkage as _lkg, fcluster as _fcl
from scipy.spatial.distance import squareform as _sqf
FX = _l2n(_TV(analyzer="char_wb", ngram_range=(4, 4), max_features=200000,
              sublinear_tf=True).fit_transform(frag_body))
FS = (FX @ FX.T).toarray()
FD = np.clip(1 - FS, 0, None); np.fill_diagonal(FD, 0)
flab = _fcl(_lkg(_sqf(FD, checks=False), method="average"), 1 - CORE_CUT,
            criterion="distance")
clusters = defaultdict(list)
for fi, l in enumerate(flab): clusters[l].append(fi)
core, small = [], []
for members in clusters.values():
    ndocs = sum(len(frags[fi]) for fi in members)
    (core if ndocs >= MIN_CORE_DOCS else small).append(members)
core.sort(key=lambda m: -sum(len(frags[fi]) for fi in m))
frag_attach = {}                   # frag idx -> (core_idx or None, confidence)
for ci, members in enumerate(core):
    for fi in members: frag_attach[fi] = (ci, 1.0)
core_fis = [fi for m in core for fi in m]
for members in small:
    for fi in members:
        sims = FS[fi, core_fis]
        j = int(np.argmax(sims)) if len(core_fis) else 0
        if len(core_fis) and sims[j] >= ATTACH_MIN:
            frag_attach[fi] = (frag_attach[core_fis[j]][0], float(sims[j]))
        else:
            frag_attach[fi] = (None, float(sims[j]) if len(core_fis) else 0.0)

core_groups, core_conf = [], []    # true versions
for ci in range(len(core)):
    mem, conf = [], {}
    for fi, (c, s) in frag_attach.items():
        if c == ci:
            for di in frags[fi]:
                mem.append(di); conf[di] = s
    core_groups.append(mem); core_conf.append(conf)
review_groups = []                 # readable but matched nothing: NOT versions
for fi, (c, s) in sorted(frag_attach.items()):
    if c is None:
        review_groups.append(list(frags[fi]))
order = sorted(range(len(core_groups)),
               key=lambda g: (-len(core_groups[g]),
                              min(docs[d]["covid"] for d in core_groups[g])))
core_groups = [core_groups[g] for g in order]
core_conf = [core_conf[g] for g in order]
review_groups.sort(key=lambda m: min(docs[d]["covid"] for d in m))
print(f"{len(core_groups)} versions, {sum(map(len,review_groups))} docs for review  {time.time()-t0:.0f}s")

# unparsed docs: for review, flagged
unparsed = [di for di, d in enumerate(docs) if not d["parsed"]]

versions = []   # dicts; status: version | review_nomatch | review_unparsed
doc_version = {}
doc_conf = {}
def rep_key(di):
    """prefer reps whose section 1 starts near the body start (no missing head),
    with a full set of sections and the longest text"""
    d = docs[di]
    early = 1 if d.get("start_off", 10**9) < 4000 else 0
    return (early, min(len(doc_secmap[di]), 34), len(d["body"]))
for gi, members in enumerate(core_groups, 1):
    vid = f"V{gi:02d}"
    rep = max(members, key=rep_key)
    for di in members:
        doc_version[di] = vid
        doc_conf[di] = core_conf[gi - 1].get(di, 1.0)
    versions.append({"vid": vid, "members": sorted(members, key=lambda x: docs[x]["covid"]),
                     "rep": rep, "status": "version", "unparsed": False})
rk = 0
for members in review_groups:
    rk += 1
    vid = f"R{rk:02d}"
    rep = max(members, key=rep_key)
    for di in members: doc_version[di] = vid
    versions.append({"vid": vid, "members": sorted(members, key=lambda x: docs[x]["covid"]),
                     "rep": rep, "status": "review_nomatch", "unparsed": False})
for k, di in enumerate(sorted(unparsed, key=lambda x: docs[x]["covid"]), 1):
    vid = f"U{k:02d}"
    doc_version[di] = vid
    versions.append({"vid": vid, "members": [di], "rep": di,
                     "status": "review_unparsed", "unparsed": True})
groups = core_groups   # for stats
STATUS_LABEL = {"version": "Version",
                "review_nomatch": "FOR REVIEW - matched no version (likely bad scan)",
                "review_unparsed": "FOR REVIEW - unreadable scan (no sections parsed)"}

def clean_sec(t):
    """final cosmetic cleanup of exported section text: page breaks, residual
    footer lines, blank-line runs"""
    t = t.replace("\x0c", "\n")
    t = scrub_footers(t)
    t = re.sub(r"(?mi)^.{0,40}Doc\.?\s*ID:?\s*[\d\-]+.*$", "", t)
    t = re.sub(r"(?m)^[\W_]{1,6}$", "", t)      # lines of pure punctuation
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()

# version-level variant per slot = majority vote across member docs
ver_secs = {}                      # vid -> [(secno, sid, vn, text)] from rep, vote-corrected
for v in versions:
    if v["unparsed"]:
        ver_secs[v["vid"]] = []
        continue
    votes = defaultdict(Counter)   # sid -> Counter of vn
    for di in v["members"]:
        for _, sid, vn, _ in doc_secmap[di]:
            votes[sid][vn] += 1
    # for display text, use the MEDIAN-length copy among this version's own member
    # docs carrying the modal variant (robust to both truncated and junk-appended copies)
    member_secs = defaultdict(list)     # (sid, vn) -> [sec idx]
    for di in v["members"]:
        for _, sid, vn, si in doc_secmap[di]:
            member_secs[(sid, vn)].append(si)
    out = []
    for no, sid, vn, si in sorted(doc_secmap[v["rep"]], key=lambda x: x[0]):
        vn2 = votes[sid].most_common(1)[0][0]
        cands = member_secs.get((sid, vn2), [])
        if cands:
            cands = sorted(cands, key=lambda i: len(sections[i]))
            rep_i = cands[len(cands) // 2]
        else:
            rep_i = variants[(sid, vn2)]["rep"]
        out.append((no, sid, vn2, clean_sec(sections[rep_i])))
    ver_secs[v["vid"]] = out

# interest rate on unpaid fees, per version — from the LIEN section:
# "...sums due ... shall bear interest at the lesser of the maximum non-usurious
#  lawful rate allowed by law or 18 percent per year"
INT_WORDS = {"eighteen": "18", "fifteen": "15", "twelve": "12", "ten": "10",
             "eight": "8", "six": "6"}
INT_RX = re.compile(
    r"bear\s+interest.{0,220}?"
    r"(?:([\d\[\{lI|OQ]{1,3}(?:\.\d+)?)\s*(?:%|percent)"
    r"|(eighteen|fifteen|twelve|ten|eight|six)\s+percent)", re.I | re.S)
def fix_digits(s):
    """repair OCR-mangled digits: '[8' -> 18, 'l8' -> 18, '1O' -> 10"""
    return s.translate(str.maketrans("[{lI|OQ", "1111100"))
unpaid_rate, unpaid_src = {}, {}
for v in versions:
    rate, src = "", ""
    for no, sid, vn, t in ver_secs[v["vid"]]:
        m = INT_RX.search(t)
        if m:
            val = fix_digits(m.group(1)) if m.group(1) else INT_WORDS.get(
                m.group(2).lower(), m.group(2))
            rate, src = val + "%", f"§{no} {topic_of_slot[sid]}"
            break
    unpaid_rate[v["vid"]] = rate
    unpaid_src[v["vid"]] = src

variant_versions = defaultdict(set)
for v in versions:
    for no, sid, vn, _ in ver_secs[v["vid"]]:
        variant_versions[(sid, vn)].add(v["vid"])

# topic-level variant numbering. Variants whose TEMPLATE language matches once
# fill-in content (names, addresses, allocations) is masked share ONE number:
# the beneficiary list differing between documents does not count as different
# language. v1 = language used by the most versions.
tvn_of = {}
by_topic = defaultdict(list)
for k in variant_versions:
    by_topic[topic_of_slot[k[0]]].append(k)
def same_language(ma, mb):
    """two-signal test on fill-in-masked text. The word-set overlap (>= 0.94) is
    the main discriminator: real template changes add/remove phrases (set 0.88 and
    below), fill-in differences leave the word-set nearly intact (0.95+). The
    sequence bar (>= 0.90) guards against same-words-reordered rewordings."""
    sm = SequenceMatcher(None, ma[:2500], mb[:2500])
    if sm.quick_ratio() < 0.90 or sm.ratio() < 0.90:
        return False
    ca, cb = Counter(ma.split()), Counter(mb.split())
    union = sum((ca | cb).values())
    return union == 0 or sum((ca & cb).values()) / union >= 0.94
def variant_median_sec(k):
    mem = sorted(variants[k]["members"], key=lambda i: len(sections[i]))
    return mem[len(mem) // 2]
for tp, ks in by_topic.items():
    ks_sorted = sorted(ks, key=lambda k: (-len(variant_versions[k]), k))
    classes = []                     # [(masked_rep_text, [variant keys])]
    for k in ks_sorted:
        mk = masked(variant_median_sec(k))
        placed = False
        for crep, mem in classes:
            if same_language(mk, crep):
                mem.append(k); placed = True
                break
        if not placed:
            classes.append((mk, [k]))
    classes.sort(key=lambda c: -sum(len(variant_versions[k]) for k in c[1]))
    for n, (crep, mem) in enumerate(classes, 1):
        for k in mem:
            tvn_of[k] = n
if os.environ.get("DBG_PAIR"):
    sa, sb = os.environ["DBG_PAIR"].split(",")
    ka = tuple([sa.split("_")[0], int(sa.split("_v")[1])])
    kb = tuple([sb.split("_")[0], int(sb.split("_v")[1])])
    ma, mb = masked(variant_median_sec(ka)), masked(variant_median_sec(kb))
    sm = SequenceMatcher(None, ma[:2500], mb[:2500])
    ca, cb = Counter(ma.split()), Counter(mb.split())
    st = sum((ca & cb).values()) / max(sum((ca | cb).values()), 1)
    print(f"DBG {sa} vs {sb}: seq={sm.ratio():.3f} quick={sm.quick_ratio():.3f} set={st:.3f}")
    print(f"DBG tvn: {tvn_of.get(ka)} vs {tvn_of.get(kb)} topic {topic_of_slot[ka[0]]} / {topic_of_slot[kb[0]]}")
    wa, wb = ma.split(), mb.split()
    smw = SequenceMatcher(None, wa, wb)
    for tag, i1, i2, j1, j2 in smw.get_opcodes():
        if tag != "equal":
            print(f"DBG [{tag}] A({i2-i1}w): {' '.join(wa[i1:i2])[:170]}")
            print(f"DBG          B({j2-j1}w): {' '.join(wb[j1:j2])[:170]}")

# ============================================================ 5. outputs
os.makedirs(OUT, exist_ok=True)
os.makedirs(os.path.join(OUT, "versions"), exist_ok=True)
def xclean(s): return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", s)
CHUNK = 30000
today = datetime.date.today().isoformat()

issues = []
cc = Counter(d["covid"] for d in docs)
issues += [(c, f"{k} PDFs share this CovID") for c, k in cc.items() if k > 1]
issues += [(docs[di]["covid"], "sections could not be parsed (bad scan) - full text kept, compared as a whole")
           for di in unparsed]
for d in docs:
    if len(d["body"]) < 2000:
        issues.append((d["covid"], f"very little text ({len(d['body'])} chars) - likely plat map/handwriting"))
for di, c in doc_conf.items():
    if c < 0.75:
        issues.append((docs[di]["covid"],
                       f"low-confidence version match ({c:.2f}) to {doc_version[di]} - verify manually"))
for d in docs:
    if d["parsed"] and d.get("start_off", 0) > 4500:
        issues.append((d["covid"],
                       f"section 1 found {d['start_off']} chars into the document - text before it may be missing from sections"))
for v in versions:
    if v["status"] == "version" and docs[v["rep"]].get("start_off", 0) > 4500:
        issues.append((docs[v["rep"]]["covid"],
                       f"representative doc for {v['vid']} may be missing text at start - review"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
wb = Workbook()
boldf = Font(name="Arial", bold=True, color="FFFFFF")
fillf = PatternFill("solid", fgColor="1F4E79")
arial = Font(name="Arial")
def style(ws, widths=None):
    for r in ws.iter_rows():
        for c in r: c.font = arial
    for c in ws[1]: c.font = boldf; c.fill = fillf
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

# --- Versions
ws = wb.active; ws.title = "Versions"
vtexts = {}
for v in versions:
    if v["unparsed"]:
        vtexts[v["vid"]] = xclean(clean_sec(docs[v["rep"]]["body"]))
    else:
        vtexts[v["vid"]] = xclean("\n\n".join(f"{no}. {t}" if not re.match(r"^\W{0,3}\d", t) else t
                                              for no, _, _, t in ver_secs[v["vid"]]))
max_chunks = max(len(t) // CHUNK + 1 for t in vtexts.values())
ws.append(["VersionID", "Status", "DocCount", "SectionCount", "SampleCovID",
           "UnpaidInterestPercent", "UnpaidInterestSource"]
          + [f"VersionText_{i+1}" for i in range(max_chunks)])
for v in versions:
    t = vtexts[v["vid"]]
    chunks = [t[i:i+CHUNK] for i in range(0, len(t), CHUNK)] or [""]
    ws.append([v["vid"], STATUS_LABEL[v["status"]], len(v["members"]),
               len(ver_secs[v["vid"]]), docs[v["rep"]]["covid"],
               unpaid_rate[v["vid"]], unpaid_src[v["vid"]]]
              + chunks + [""] * (max_chunks - len(chunks)))
style(ws, [10, 40, 9, 12, 11, 16, 22])

# --- CovID_Map
ws2 = wb.create_sheet("CovID_Map")
ws2.append(["CovID", "VersionID", "MatchConfidence", "Filename", "RelativePath",
            "Pages", "OCR_Used", "SectionsParsed"])
for di in sorted(range(len(docs)), key=lambda x: (docs[x]["covid"], docs[x]["filename"])):
    d = docs[di]
    ws2.append([d["covid"], doc_version[di], round(doc_conf.get(di, 1.0), 3),
                d["filename"], d["relpath"], d["pages"],
                "Yes" if d["ocr"] else "No", len(d["secs"])])
style(ws2, [10, 10, 15, 30, 40, 8, 10, 13])

# --- VersionSection_Map (long: the FileMaker find table)
ws3 = wb.create_sheet("VersionSection_Map")
ws3.append(["VersionID", "SectionNo", "Topic", "TopicVariantNo", "SlotID", "VariantID",
            "RawHeading"])
def heading(t):
    h = re.sub(r"^\W{0,6}\d{0,2}\s*[\.\),:]?\s*", "", t)[:60]
    return h.split("\n")[0][:60]
for v in versions:
    for no, sid, vn, t in ver_secs[v["vid"]]:
        ws3.append([v["vid"], no, topic_of_slot[sid], tvn_of[(sid, vn)],
                    sid, f"{sid}_v{vn}", xclean(heading(t))])
style(ws3, [10, 10, 32, 14, 9, 12, 45])

# --- Section_Variants
ws4 = wb.create_sheet("Section_Variants")
ws4.append(["Topic", "TopicVariantNo", "TypicalSectionNo", "SlotID", "VariantID",
            "UsedByVersionCount", "SimilarityToV1", "UsedByVersions", "Text"])
v1rep = {sid: d["rep"] for (sid, vn), d in variants.items() if vn == 1}
for k in sorted(variant_versions, key=lambda k: (topic_no[topic_of_slot[k[0]]],
                                                 topic_of_slot[k[0]], tvn_of[k])):
    sid, vn = k
    used = sorted(variant_versions[k])
    rep_i = variants[k]["rep"]
    s2 = 1.0 if vn == 1 else round(SequenceMatcher(None, snorms[rep_i][:1500],
                                                   snorms[v1rep[sid]][:1500]).ratio(), 3)
    ws4.append([topic_of_slot[sid], tvn_of[k], topic_no[topic_of_slot[sid]], sid,
                f"{sid}_v{vn}", len(used), s2,
                ", ".join(used[:150]), xclean(sections[rep_i])[:32000]])
style(ws4, [32, 14, 15, 9, 12, 17, 13, 40, 80])

# --- Topics_Matrix (rows = topics, columns = versions; cell = section number + variant)
ws6 = wb.create_sheet("Topics_Matrix")
true_vs = [v for v in versions if v["status"] == "version"]
ver_topic = {}                    # (vid, topic) -> "§6 v2"
for v in versions:
    for no, sid, vn, _ in ver_secs[v["vid"]]:
        tp = topic_of_slot[sid]
        cell = f"§{no} v{tvn_of[(sid, vn)]}"
        key = (v["vid"], tp)
        ver_topic[key] = cell if key not in ver_topic else ver_topic[key] + " + " + cell
true_ids = {v["vid"] for v in true_vs}
topic_rows = sorted({tp for (vid, tp) in ver_topic if vid in true_ids},
                    key=lambda tp: (topic_no[tp], tp))
ws6.append(["Topic", "TypicalSectionNo", "VariantCount"] + [v["vid"] for v in true_vs])
for tp in topic_rows:
    nvar = len({tvn_of[k] for k in variant_versions if topic_of_slot[k[0]] == tp})
    ws6.append([tp, topic_no[tp], nvar]
               + [ver_topic.get((v["vid"], tp), "MISSING") for v in true_vs])
style(ws6, [32, 15, 12] + [9] * len(true_vs))

# --- Versions_Sections (wide: one column per section number)
ws5 = wb.create_sheet("Versions_Sections")
max_no = max((no for v in versions for no, *_ in ver_secs[v["vid"]]), default=0)
ws5.append(["VersionID", "DocCount"] + [f"Sec_{i}" for i in range(1, max_no + 1)])
for v in versions:
    row = [""] * max_no
    for no, sid, vn, _ in ver_secs[v["vid"]]:
        tag = f"{sid}_v{vn}"
        row[no - 1] = tag if not row[no - 1] else row[no - 1] + "+" + tag
    ws5.append([v["vid"], len(v["members"])] + row)
style(ws5, [10, 9] + [13] * max_no)
wb.save(os.path.join(OUT, "covenant_matrix.xlsx"))
print(f"xlsx  {time.time()-t0:.0f}s")

# --- versions/*.txt + issues
for v in versions:
    with open(os.path.join(OUT, "versions", f"{v['vid']}.txt"), "w", encoding="utf-8") as f:
        f.write(vtexts[v["vid"]])
with open(os.path.join(OUT, "issues.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(["CovID", "Issue"]); w.writerows(sorted(set(issues)))

# --- json
json.dump({"stats": {"docs": len(docs), "versions": len(groups),
                     "review": sum(1 for v in versions if v["status"] != "version"),
                     "slots": len(slot_members), "variants": len(variants), "date": today},
           "versions": [{"version_id": v["vid"], "status": STATUS_LABEL[v["status"]],
                         "doc_count": len(v["members"]),
                         "unpaid_interest_percent": unpaid_rate[v["vid"]],
                         "unpaid_interest_source": unpaid_src[v["vid"]],
                         "sample_covid": docs[v["rep"]]["covid"],
                         "covids": [docs[di]["covid"] for di in v["members"]],
                         "sections": [{"no": no, "topic": topic_of_slot[sid],
                                       "topic_variant": tvn_of[(sid, vn)],
                                       "slot": sid, "variant": f"{sid}_v{vn}",
                                       "text": t} for no, sid, vn, t in ver_secs[v["vid"]]],
                         "unparsed_text": vtexts[v["vid"]] if v["unparsed"] else ""}
                        for v in versions],
           "covid_map": [{"covid": docs[di]["covid"], "version_id": doc_version[di],
                          "confidence": round(doc_conf.get(di, 1.0), 3),
                          "filename": docs[di]["filename"], "relpath": docs[di]["relpath"]}
                         for di in range(len(docs))],
           "section_variants": [{"topic": topic_of_slot[sid],
                                 "topic_variant": tvn_of[(sid, vn)],
                                 "slot": sid, "typical_section_no": slot_no[sid],
                                 "variant": f"{sid}_v{vn}",
                                 "used_by": sorted(variant_versions[(sid, vn)]),
                                 "text": sections[variants[(sid, vn)]["rep"]]}
                                for (sid, vn) in sorted(variant_versions,
                                                        key=lambda k: (slot_no[k[0]], k[0], k[1]))]},
          open(os.path.join(OUT, "covenant_matrix.json"), "w", encoding="utf-8"), indent=1)

# --- html
topic_variants_h = defaultdict(list)
for k in sorted(variant_versions, key=lambda k: (topic_of_slot[k[0]], tvn_of[k])):
    sid, vn = k
    topic_variants_h[topic_of_slot[sid]].append(
        [f"{sid}_v{vn}", tvn_of[k], sorted(variant_versions[k]),
         sections[variants[k]["rep"]]])
hdata = {"stats": {"docs": len(docs), "versions": len(groups),
                   "review": sum(1 for v in versions if v["status"] != "version"),
                   "topics": len({topic_of_slot[k[0]] for k in variant_versions}),
                   "variants": len(variant_versions), "date": today},
         "topicVariants": topic_variants_h,
         "topicNo": {tp: topic_no[tp] for tp in topic_variants_h},
         "versions": [{"id": v["vid"], "docCount": len(v["members"]),
                       "status": v["status"],
                       "rate": unpaid_rate[v["vid"]], "rateSrc": unpaid_src[v["vid"]],
                       "sample": docs[v["rep"]]["covid"],
                       "covids": [docs[di]["covid"] for di in v["members"]],
                       "secs": [[no, f"{sid}_v{vn}", t, topic_of_slot[sid], tvn_of[(sid, vn)]]
                                for no, sid, vn, t in ver_secs[v["vid"]]],
                       "raw": vtexts[v["vid"]] if v["unparsed"] else ""}
                      for v in versions]}
payload = json.dumps(hdata).replace("</", "<\\/")
tpl = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Covenant Version Matrix — sections</title>
<style>
 :root{--blue:#1F4E79;--lt:#eef3f8;--hl:#ffe066;--add:#d4f7d4;--del:#ffd9d9}
 *{box-sizing:border-box}
 body{font-family:-apple-system,Segoe UI,Arial,sans-serif;margin:0;background:#f5f6f8;color:#222}
 header{background:var(--blue);color:#fff;padding:14px 28px}
 header h1{margin:0 0 6px;font-size:20px}
 .stats{display:flex;gap:24px;font-size:12.5px;opacity:.92;flex-wrap:wrap}
 .stats b{font-size:16px;display:block}
 .bar{position:sticky;top:0;background:#fff;padding:10px 28px;box-shadow:0 1px 4px rgba(0,0,0,.12);z-index:5;display:flex;gap:12px;align-items:center;flex-wrap:wrap}
 #q{flex:1;min-width:240px;max-width:520px;padding:8px 13px;font-size:14.5px;border:1.5px solid #bbb;border-radius:8px}
 #hits{font-size:13px;color:#555}
 label.f{font-size:13px;color:#444;display:flex;gap:5px;align-items:center}
 #cmpbtn,#secbtn,#helpbtn{padding:7px 14px;border:1.5px solid var(--blue);background:#fff;color:var(--blue);border-radius:8px;cursor:pointer;font-size:13.5px}
 #cmpbtn:disabled{opacity:.4;cursor:default}
 #helpbtn{font-weight:700;padding:7px 12px}
 #hintbar{background:#fdf6dd;border-bottom:1px solid #eadfa8;padding:7px 28px;font-size:12.5px;color:#5a4d1a}
 #hintbar a{color:var(--blue)}
 .toprow{display:flex;justify-content:space-between;gap:10px;padding:8px 0;border-bottom:1px solid #eef1f5;font-size:13.5px;cursor:pointer;align-items:center}
 .toprow:hover{background:var(--lt)}
 .helpsec{margin:0 0 18px}
 .helpsec h3{margin:0 0 6px;font-size:14.5px;color:var(--blue)}
 .helpsec p{margin:0 0 6px;font-size:13.5px;line-height:1.55}
 kbd{background:var(--lt);border:1px solid #c9d7e6;border-radius:4px;padding:0 5px;font-size:12px}
 main{padding:16px 28px;max-width:1250px;margin:0 auto}
 .ver{background:#fff;border:1px solid #dde3ea;border-radius:10px;margin-bottom:11px;overflow:hidden}
 .vhead{display:flex;align-items:center;gap:12px;padding:10px 15px;cursor:pointer;background:var(--lt)}
 .vhead:hover{background:#e2ebf5}
 .vid{font-weight:700;color:var(--blue);font-size:16px;min-width:52px}
 .meta{font-size:12.5px;color:#555}
 .cmp{margin-left:auto;font-size:12px;display:flex;gap:5px;align-items:center;color:#444}
 .chips{padding:7px 15px;border-top:1px solid #e8edf2;display:flex;flex-wrap:wrap;gap:4px;max-height:100px;overflow:auto}
 .chip{background:var(--lt);border:1px solid #c9d7e6;border-radius:12px;padding:1px 8px;font-size:11.5px;font-family:ui-monospace,monospace}
 .chip.hit{background:var(--hl);border-color:#d4b106}
 .body{display:none;border-top:1px solid #e8edf2;padding:11px 15px;max-height:520px;overflow:auto}
 .ver.open .body{display:block}
 .para{display:flex;gap:10px;padding:5px 0;font-size:13.5px;line-height:1.5;border-bottom:1px dotted #f0f2f5}
 .pn{color:#fff;background:#8fa8c2;border-radius:4px;font-family:ui-monospace,monospace;font-size:10.5px;min-width:90px;max-width:185px;text-align:center;flex-shrink:0;align-self:flex-start;padding:2px 5px}
 .pn.alt{background:#c0392b}
 .pn.vhi{outline:2px solid #c0392b}
 mark{background:var(--hl);padding:0 1px}
 .arrow{transition:.2s;color:#888}
 .ver.open .arrow{transform:rotate(90deg)}
 .hidden{display:none}
 #overlay{display:none;position:fixed;inset:0;background:rgba(20,30,45,.55);z-index:20}
 #panel{position:fixed;inset:3% 4%;background:#fff;border-radius:12px;z-index:21;display:none;flex-direction:column;overflow:hidden}
 #phead{background:var(--blue);color:#fff;padding:12px 20px;display:flex;gap:16px;align-items:center}
 #phead b{font-size:16px}
 #pclose{margin-left:auto;background:none;border:1px solid #fff;color:#fff;border-radius:6px;padding:4px 12px;cursor:pointer}
 #pbody{overflow:auto;padding:14px 20px;flex:1}
 .drow{display:grid;grid-template-columns:100px 1fr 1fr;gap:10px;border-bottom:1px solid #eef1f5;padding:7px 0;font-size:13px;line-height:1.5}
 .drow.same{color:#999;font-size:12px}
 .drow .slot{font-family:ui-monospace,monospace;font-size:11px;color:#555}
 .drow ins{background:var(--add);text-decoration:none}
 .drow del{background:var(--del);text-decoration:none}
 .dhead{position:sticky;top:0;background:#fff;font-weight:700;border-bottom:2px solid var(--blue)}
 label.g{font-size:12.5px;color:#eee;display:flex;gap:5px;align-items:center}
</style></head><body>
<header><h1>Covenant Version Matrix — numbered sections only</h1>
<div class="stats">
 <span><b id=s1></b>documents</span><span><b id=s2></b>versions</span>
 <span><b id=s3></b>docs for review (not versions)</span><span><b id=s4></b>section topics</span>
 <span><b id=s5></b>language variants</span><span><b id=s6></b>generated</span>
</div></header>
<div class="bar"><input id="q" placeholder="Search a CovID (4417), a phrase (transfer fee), or a section name (exemptions)…" autofocus>
<button id="secbtn" title="Browse every section topic and its language variants">Browse sections</button>
<label class="f"><input type="checkbox" id="multi" checked> hide one-off versions</label>
<button id="cmpbtn" disabled>Compare (pick 2)</button>
<button id="helpbtn" title="How to use this page">?</button>
<span id="hits"></span></div>
<div id="hintbar">Tip: click any blue <b>§-badge</b> inside a version to see every wording of that paragraph across all versions · tick <b>compare</b> on two versions for a side-by-side · <a href="#" id="hintmore">full guide</a></div>
<main id="list"></main>
<div id="overlay"></div>
<div id="panel"><div id="phead"><b id=ptitle></b>
<label class="g" id="hidesamewrap"><input type="checkbox" id="hidesame" checked> hide identical sections</label>
<label class="g" id="fulltxwrap" style="display:none"><input type="checkbox" id="fulltx"> show full text of every variant</label>
<button id="pclose">Close</button></div><div id="pbody"></div></div>
<script>
const DATA = __DATA__;
const esc=s=>s.replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));
const reviewN=DATA.versions.filter(v=>v.status!=='version').reduce((a,v)=>a+v.docCount,0);
s1.textContent=DATA.stats.docs;s2.textContent=DATA.stats.versions;
s3.textContent=reviewN;s4.textContent=DATA.stats.topics;
s5.textContent=DATA.stats.variants;s6.textContent=DATA.stats.date;
const STATUS={review_nomatch:'FOR REVIEW — matched no version (likely bad scan)',
              review_unparsed:'FOR REVIEW — unreadable scan'};
const list=document.getElementById('list'),els={},sel=new Set();
let divDone=false;
DATA.versions.forEach(v=>{
 const isRev=v.status!=='version';
 if(isRev&&!divDone){
  divDone=true;
  const h=document.createElement('div');
  h.innerHTML=`<h2 style="color:#c0392b;font-size:16px;border-bottom:2px solid #c0392b;padding-bottom:6px;margin:26px 0 12px">For review — not counted as versions (bad/unreadable scans)</h2>`;
  list.appendChild(h);
 }
 const d=document.createElement('div');d.className='ver';d.dataset.id=v.id;
 const un=v.id[0]==='U';
 d.innerHTML=`<div class="vhead"><span class="vid"${isRev?' style="color:#c0392b"':''}>${v.id}</span>
  <span class="meta">${v.docCount} doc${v.docCount!==1?'s':''} · ${un?'unreadable scan':v.secs.length+' sections'} · sample ${v.sample}${v.rate?` · <b title="interest rate on unpaid fees (${esc(v.rateSrc)})">unpaid interest ${v.rate}</b>`:(v.status==='version'?' · <span title="no unpaid-fee interest clause found in this template">unpaid interest —</span>':'')}${isRev?' · <b style="color:#c0392b">'+STATUS[v.status]+'</b>':''}</span>
  <label class="cmp" onclick="event.stopPropagation()"><input type="checkbox" data-v="${v.id}" class="selbox">compare</label>
  <span class="arrow">▶</span></div>
  <div class="chips">${v.covids.slice(0,400).map(c=>`<span class="chip" data-c="${c}">${c}</span>`).join('')}</div>
  <div class="body"></div>`;
 d.querySelector('.vhead').onclick=()=>{d.classList.toggle('open');if(d.classList.contains('open'))render(v,d,curQ());};
 d.querySelector('.body').addEventListener('click',e=>{
  const pn=e.target.closest('.pn');
  if(pn&&pn.dataset.topic)showTopic(pn.dataset.topic);
 });
 els[v.id]=d;list.appendChild(d);
});
document.querySelectorAll('.selbox').forEach(b=>b.addEventListener('change',()=>{
 if(b.checked)sel.add(b.dataset.v);else sel.delete(b.dataset.v);
 if(sel.size>2){b.checked=false;sel.delete(b.dataset.v);return;}
 const btn=document.getElementById('cmpbtn');
 btn.disabled=sel.size!==2;
 btn.textContent=sel.size===2?`Compare ${[...sel].join(' vs ')}`:'Compare (pick 2)';
}));
function curQ(){return document.getElementById('q').value.trim().toLowerCase()}
function render(v,el,s){
 const rx=s?new RegExp('('+s.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+')','gi'):null;
 let out=[];
 if(v.raw){
  let h=esc(v.raw);if(rx)h=h.replace(rx,'<mark>$1</mark>');
  out.push(`<div class="para"><span class="pn">raw text</span><span style="white-space:pre-wrap">${h}</span></div>`);
 }
 v.secs.forEach(([no,tag,t,topic,tvn])=>{
  const tagHit=s&&(tag.toLowerCase().includes(s)||topic.toLowerCase().includes(s));
  if(s&&!t.toLowerCase().includes(s)&&!tagHit)return;
  let h=esc(t);if(rx&&!tagHit)h=h.replace(rx,'<mark>$1</mark>');
  const alt=tvn>1;
  const nvar=(DATA.topicVariants[topic]||[]).length;
  const short=topic.length>20?topic.slice(0,19)+'…':topic;
  out.push(`<div class="para"><span class="pn${alt?' alt':''}${tagHit?' vhi':''}" data-topic="${esc(topic)}" style="cursor:pointer" title="${esc(topic)} — variant v${tvn} of ${nvar} (${tag}). Click to see every language variant of this section across versions. Red = not the most common language.">§${no} ${short} v${tvn}</span><span>${h}</span></div>`);
 });
 if(s&&!out.length)out.push('<div class="para"><span></span><span style="color:#999">CovID match only</span></div>');
 el.querySelector('.body').innerHTML=out.join('');
}
const q=document.getElementById('q'),hits=document.getElementById('hits'),multi=document.getElementById('multi');
let t;q.addEventListener('input',()=>{clearTimeout(t);t=setTimeout(search,250)});
multi.addEventListener('change',search);
function search(){
 const s=curQ();let shown=0,pm=0;
 DATA.versions.forEach(v=>{
  const el=els[v.id];
  el.querySelectorAll('.chip.hit').forEach(c=>c.classList.remove('hit'));
  if(!s){const hide=multi.checked&&v.docCount===1;
   el.classList.toggle('hidden',hide);el.classList.remove('open');
   if(!hide)shown++;return;}
  const cov=v.covids.filter(c=>c.includes(s));
  let tp=0;
  v.secs.forEach(([no,tag,t,topic])=>{if(t.toLowerCase().includes(s)||tag.toLowerCase().includes(s)||topic.toLowerCase().includes(s)){tp++;pm++;}});
  if(v.raw&&v.raw.toLowerCase().includes(s))tp++;
  const match=tp>0||cov.length>0;
  el.classList.toggle('hidden',!match);
  if(match){shown++;el.classList.add('open');render(v,el,s);
   cov.forEach(c=>el.querySelector(`.chip[data-c="${c}"]`)?.classList.add('hit'));}
 });
 hits.textContent=s?`${shown} versions, ${pm} matching sections`:'';
}
search();
function wdiff(a,b){
 const A=a.split(/\s+/),B=b.split(/\s+/),n=A.length,m=B.length;
 if(n*m>4000000)return null;
 const L=Array.from({length:n+1},()=>new Int32Array(m+1));
 for(let i=n-1;i>=0;i--)for(let j=m-1;j>=0;j--)
  L[i][j]=A[i]===B[j]?L[i+1][j+1]+1:Math.max(L[i+1][j],L[i][j+1]);
 const out=[];let i=0,j=0;
 while(i<n&&j<m){
  if(A[i]===B[j]){out.push(['eq',A[i]]);i++;j++;}
  else if(L[i+1][j]>=L[i][j+1]){out.push(['del',A[i]]);i++;}
  else{out.push(['ins',B[j]]);j++;}
 }
 while(i<n)out.push(['del',A[i++]]);
 while(j<m)out.push(['ins',B[j++]]);
 return out;
}
function renderDiff(a,b){
 const ops=wdiff(a,b);
 if(!ops)return [esc(a),esc(b)];
 let la=[],lb=[];
 ops.forEach(([t,w])=>{
  if(t==='eq'){la.push(esc(w));lb.push(esc(w));}
  else if(t==='del')la.push('<del>'+esc(w)+'</del>');
  else lb.push('<ins>'+esc(w)+'</ins>');
 });
 return [la.join(' '),lb.join(' ')];
}
document.getElementById('cmpbtn').onclick=()=>{
 const [ida,idb]=[...sel];
 const va=DATA.versions.find(v=>v.id===ida),vb=DATA.versions.find(v=>v.id===idb);
 const ma={},mb={};
 va.secs.forEach(s=>{(ma[s[3]]=ma[s[3]]||[]).push(s);});
 vb.secs.forEach(s=>{(mb[s[3]]=mb[s[3]]||[]).push(s);});
 const keys=[...new Set([...va.secs,...vb.secs].map(s=>s[3]))];
 // pass 1: pair by slot; leftovers collected for pass 2
 const pairs=[],leftA=[],leftB=[];
 keys.forEach(k=>{
  const ia=ma[k]||[],ib=mb[k]||[];
  const n=Math.min(ia.length,ib.length);
  for(let x=0;x<n;x++)pairs.push([ia[x],ib[x]]);
  for(let x=n;x<ia.length;x++)leftA.push(ia[x]);
  for(let x=n;x<ib.length;x++)leftB.push(ib[x]);
 });
 // pass 2: same section NUMBER but different slot = same section, different language
 leftA.forEach(sa=>{
  const j=leftB.findIndex(sb=>sb[0]===sa[0]);
  if(j>=0){pairs.push([sa,leftB[j]]);leftB.splice(j,1);}
  else pairs.push([sa,null]);
 });
 leftB.forEach(sb=>pairs.push([null,sb]));
 pairs.sort((p,q)=>((p[0]||p[1])[0])-((q[0]||q[1])[0]));
 let rows=[`<div class="drow dhead"><span class="slot">Section</span><span>${va.id}</span><span>${vb.id}</span></div>`];
 let same=0,diff=0;
 pairs.forEach(([sa,sb])=>{
  const topic=(sa||sb)[3];
  const nums=(sa?('§'+sa[0]):'')+(sb&&(!sa||sb[0]!==sa[0])?((sa?' / ':'')+'§'+sb[0]):'');
  const lbl=`${nums}<br><b>${esc(topic)}</b>`;
  if(sa&&sb&&sa[1]===sb[1]){same++;
   rows.push(`<div class="drow same hideable"><span class="slot">${lbl}</span><span>identical language (v${sa[4]})</span><span></span></div>`);
  }else if(sa&&sb&&sa[3]===sb[3]&&sa[4]===sb[4]){same++;
   rows.push(`<div class="drow same hideable"><span class="slot">${lbl}</span><span>same template language (v${sa[4]}) — only fill-in details (names, addresses, allocations) differ</span><span></span></div>`);
  }else{diff++;
   const [ha,hb]=sa&&sb?renderDiff(sa[2],sb[2]):[sa?esc(sa[2]):'<i style="color:#bbb">not present</i>',sb?esc(sb[2]):'<i style="color:#bbb">not present</i>'];
   rows.push(`<div class="drow"><span class="slot">${lbl}<br>${sa?'v'+sa[4]:''}${sb&&(!sa||sb[4]!==sa[4])?' vs v'+sb[4]:''}</span><span>${ha}</span><span>${hb}</span></div>`);
  }
 });
 document.getElementById('hidesamewrap').style.display='flex';
 document.getElementById('hidesamewrap').style.visibility='visible';
 document.getElementById('fulltxwrap').style.display='none';
 document.getElementById('ptitle').textContent=`${va.id} vs ${vb.id} — ${diff} differing, ${same} identical sections`;
 document.getElementById('pbody').innerHTML=rows.join('');
 const hs=document.getElementById('hidesame');
 const apply=()=>document.querySelectorAll('.hideable').forEach(e=>e.style.display=hs.checked?'none':'grid');
 hs.onchange=apply;apply();
 document.getElementById('overlay').style.display='block';
 document.getElementById('panel').style.display='flex';
};
document.getElementById('pclose').onclick=document.getElementById('overlay').onclick=()=>{
 document.getElementById('overlay').style.display='none';
 document.getElementById('panel').style.display='none';
};
// ---------- section browser: all language variants of one topic ----------
function changeHunks(base,text){
 // compact change list: [pre-context, deleted, inserted, post-context]
 const ops=wdiff(base,text);
 if(!ops)return null;
 const hunks=[];let k=0;
 while(k<ops.length){
  if(ops[k][0]==='eq'){k++;continue;}
  const del=[],ins=[];const start=k;
  while(k<ops.length&&ops[k][0]!=='eq'){(ops[k][0]==='del'?del:ins).push(ops[k][1]);k++;}
  const pre=[];for(let p=start-1;p>=0&&pre.length<4;p--){if(ops[p][0]!=='eq')break;pre.unshift(ops[p][1]);}
  const post=[];for(let p=k;p<ops.length&&post.length<4;p++){if(ops[p][0]!=='eq')break;post.push(ops[p][1]);}
  hunks.push([pre.join(' '),del.join(' '),ins.join(' '),post.join(' ')]);
 }
 return hunks;
}
function showTopic(topic){
 const vars=DATA.topicVariants[topic];
 if(!vars)return;
 // group slot-variants sharing the same topic-variant number (= same template
 // language, only fill-in details differ)
 const groups={};
 vars.forEach(([tag,tvn,used,text])=>{
  if(!groups[tvn])groups[tvn]={tvn,used:new Set(),text};
  used.forEach(u=>groups[tvn].used.add(u));
 });
 const glist=Object.values(groups).sort((a,b)=>a.tvn-b.tvn);
 const base=glist[0].text;
 const full=document.getElementById('fulltx').checked;
 let rows=[`<div class="drow dhead"><span class="slot">Variant</span><span>Used by</span><span>${full?'Language (differences vs v1 highlighted)':'What changed vs v1 (red = removed, green = added)'}</span></div>`];
 glist.forEach((g,k)=>{
  const used=[...g.used].sort();
  const chips=used.map(u=>`<span class="chip" style="cursor:pointer" onclick="closePanelAndSearch('${u}')">${u}</span>`).join(' ');
  let h;
  if(k===0){
   h=full?esc(g.text):`<div style="color:#555;font-size:12.5px;line-height:1.55">${esc(g.text)}</div>`;
  }else if(full){
   const[,hb]=renderDiff(base,g.text);h=hb;
  }else{
   const hs=changeHunks(base,g.text);
   if(hs===null){const[,hb]=renderDiff(base,g.text);h=hb;}
   else h=hs.map(([pre,del,ins,post])=>
    `<div style="padding:3px 0;border-bottom:1px dotted #eee"><span style="color:#999">…${esc(pre)} </span>${del?'<del>'+esc(del)+'</del>':''}${del&&ins?' ':''}${ins?'<ins>'+esc(ins)+'</ins>':''}<span style="color:#999"> ${esc(post)}…</span></div>`).join('')
    ||'<i style="color:#999">no template change (fill-in details only)</i>';
  }
  rows.push(`<div class="drow"><span class="slot"><b>v${g.tvn}</b>${k===0?'<br><span style="color:#999">reference</span>':''}<br>${used.length} version${used.length!==1?'s':''}</span><span style="display:flex;flex-wrap:wrap;gap:3px;align-content:flex-start">${chips}</span><span>${h}</span></div>`);
 });
 document.getElementById('hidesamewrap').style.display='none';
 document.getElementById('fulltxwrap').style.display='flex';
 document.getElementById('fulltx').onchange=()=>showTopic(topic);
 document.getElementById('ptitle').textContent=`${topic} — ${glist.length} language variant${glist.length!==1?'s':''} across versions`;
 document.getElementById('pbody').innerHTML=rows.join('');
 document.getElementById('overlay').style.display='block';
 document.getElementById('panel').style.display='flex';
}
// ---------- sections index ----------
document.getElementById('secbtn').onclick=()=>{
 const tps=Object.keys(DATA.topicVariants).sort((a,b)=>(DATA.topicNo[a]-DATA.topicNo[b])||a.localeCompare(b));
 let rows=[`<div class="drow dhead"><span class="slot">Typical §</span><span>Section</span><span>Language variants</span></div>`];
 tps.forEach(tp=>{
  const nvar=new Set(DATA.topicVariants[tp].map(v=>v[1])).size;
  const nver=new Set(DATA.topicVariants[tp].flatMap(v=>v[2])).size;
  rows.push(`<div class="toprow" onclick="showTopic(${JSON.stringify(tp).replace(/"/g,'&quot;')})"><span style="min-width:70px;color:#888">§${DATA.topicNo[tp]}</span><span style="flex:1;font-weight:600">${esc(tp)}</span><span style="color:#666">${nvar} variant${nvar!==1?'s':''} · ${nver} versions</span></div>`);
 });
 document.getElementById('hidesamewrap').style.display='none';
 document.getElementById('fulltxwrap').style.display='none';
 document.getElementById('ptitle').textContent='All sections — click one to see every wording across versions';
 document.getElementById('pbody').innerHTML=rows.map(r=>r.replace('class="drow dhead"','class="drow dhead" style="grid-template-columns:70px 1fr 220px"')).join('');
 document.getElementById('overlay').style.display='block';
 document.getElementById('panel').style.display='flex';
};
// ---------- help ----------
const HELP=`
<div class="helpsec"><h3>What this page shows</h3>
<p>Every covenant document was read (OCR), split into its numbered sections, and grouped: a <b>version</b> is a distinct combination of section language. Grantor info, notary blocks and Exhibit A legal descriptions are excluded. Names, addresses and dollar/percent figures never count as language differences.</p></div>
<div class="helpsec"><h3>Search box</h3>
<p>Type a <b>CovID</b> (e.g. <kbd>4417</kbd>) to find which version that document belongs to — its chip lights up yellow. Type a <b>phrase</b> (e.g. <kbd>transfer fee</kbd>) to see every version's matching paragraphs, highlighted. Type a <b>section name</b> (e.g. <kbd>exemptions</kbd>) to filter to that section in every version.</p></div>
<div class="helpsec"><h3>See every wording of one paragraph</h3>
<p>Use <b>Browse sections</b> (top bar) and pick a section — or click any blue <b>§-badge</b> inside an opened version. You get variant v1 in full, then each other variant reduced to just what changed: <del style="background:var(--del)">removed</del> <ins style="background:var(--add)">added</ins>. The chips show which versions use each wording; click a chip to jump there. Tick "show full text" for complete texts.</p></div>
<div class="helpsec"><h3>Compare two whole versions</h3>
<p>Tick <b>compare</b> on any two versions, then press the Compare button. Identical sections are hidden (untick to show); differing ones appear side-by-side with word-level highlights. "Same template language — only fill-in details differ" means the wording matches and only names/amounts vary.</p></div>
<div class="helpsec"><h3>Reading the badges</h3>
<p><kbd>§6 EXEMPTIONS v3</kbd> = this version's section 6 is the EXEMPTIONS paragraph, wearing language variant 3. A <b style="color:#c0392b">red badge</b> = not the most common wording of that section. Versions marked <b style="color:#c0392b">FOR REVIEW</b> are bad or unreadable scans and are not counted as versions.</p></div>`;
document.getElementById('helpbtn').onclick=document.getElementById('hintmore').onclick=(e)=>{
 e.preventDefault();
 document.getElementById('hidesamewrap').style.display='none';
 document.getElementById('fulltxwrap').style.display='none';
 document.getElementById('ptitle').textContent='How to use this page';
 document.getElementById('pbody').innerHTML=HELP;
 document.getElementById('overlay').style.display='block';
 document.getElementById('panel').style.display='flex';
};
function closePanelAndSearch(vid){
 document.getElementById('overlay').style.display='none';
 document.getElementById('panel').style.display='none';
 const el=els[vid];
 if(el){el.classList.remove('hidden');el.scrollIntoView({behavior:'smooth',block:'center'});
  el.style.outline='3px solid #c0392b';setTimeout(()=>el.style.outline='',2500);}
}
</script></body></html>"""
open(os.path.join(OUT, "covenant_matrix.html"), "w", encoding="utf-8").write(
    tpl.replace("__DATA__", payload))
print(f"ALL DONE  {time.time()-t0:.0f}s")
